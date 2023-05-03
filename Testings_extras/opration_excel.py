import openpyxl
import pandas as pd
# Open the Excel file
# path=r'C:\Users\dell7480\Desktop\mcom tools\Trend_Project\Tnch_trend\Tnch_NPN_trend_project(0.2)\output\Tnch_trend_output.xlsx'
path=r"C:\Users\dell7480\Desktop\pre_post_tool\7day kpi\3sites_pre_Post_kpi\test.xlsx"
# path=r"C:\Users\dell7480\Desktop\mcom tools\Trend_Project\Tnch_trend\Tnch_NPN_trend_project(0.2)\output\3sitepost_Tnch_trend_output_7DAY.xlsx"
# path=r"C:\Users\dell7480\Desktop\mcom tools\Trend_Project\Tnch_trend\Tnch_NPN_trend_project(0.2)\output\3sitepost_Tnch_trend_output_7DAY\.xlsx"
# workbook = openpyxl.load_workbook(path,data_only=True)
# workbook.save("values.xlsx")


# # Select the sheet you want to read
# worksheet = workbook['L900']

# # Read the values of cells with formulas
# for row in worksheet.iter_rows(min_row=4, max_row=10, min_col=1, max_col=3, values_only=True):
#     print(row)
#     print()



################################################   Code To Delete rows in a excel  based on a value in a column    ##########################################

# import openpyxl

# # Load the Excel file
# workbook = openpyxl.load_workbook(path)
# print(workbook.sheetnames)
# # Select the worksheet
# worksheet = workbook["L900"]

# # Define the column index where NaN values need to be checked
# column_index = 1

# # Get the maximum number of rows in the worksheet
# max_rows = worksheet.max_row

# # Loop through all rows in the worksheet
# for row in range(max_rows, 3, -1):
#     # Get the value in the specified column
#     value = worksheet.cell(row=row, column=column_index).value
#     print(value)
#     # Check if the value is NaN
#     if value is None :
#         # Delete the entire row
#         worksheet.delete_rows(row)

# # Save the changes to the Excel file
# workbook.save('example.xlsx')

# ########################################################******************************************##########################################
# ############################################################# to get excel row data as a list #########################
# row_data = []

# # # Iterate over the rows in the worksheet and append the cell values to the row_data list
# # for row in worksheet.iter_rows(min_row=4, max_row=4,):
# #     print(row)
# #     row_data.extend(row)

# # # Print the row data as a list
# # print(row_data)


# lis=[]

# # for row in worksheet.iter_rows(min_row=4, max_row=4):
# #     for cell in row:
# #         lis.append(cell.value)
    
# # print(lis)
# # print("done")

# row=worksheet[4]
# for cell in row:
#         lis.append(cell.value)
    
# print(lis)

# listt=["ram"]
# worksheet.append(lis)
# workbook.save("ram.xlsx")
# print("done")







workbook = openpyxl.load_workbook(path)
lis_worksheet=workbook.sheetnames
# Select the worksheet


# Define the column index where NaN values need to be checked
column_index = 1

# Get the maximum number of rows in the worksheet



# loop for all  the sheets/............
for ws in  lis_worksheet:
    worksheet=workbook[ws]
    max_rows = worksheet.max_row
    # Loop through all rows in the worksheet
    for row in range(max_rows, 3, -1):
        print(".")
        print("#.")
        # Get the value in the specified column
        value = worksheet.cell(row=row, column=column_index).value
        print(value)
        # Check if the value is NaN
        if value is None :
            # Delete the entire row
            worksheet.delete_rows(row)

# Save the changes to the Excel file
workbook.save('example.xlsx')


