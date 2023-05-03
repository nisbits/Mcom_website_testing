import openpyxl
from openpyxl.styles import PatternFill
 
wb = openpyxl.load_workbook(r"C:\Users\dell7480\Desktop\Mobile_com_web_app\mcom_website\post_trend_excel.xlsx")
ws = wb['Sheet1']
colors = ['00660066', '00FFFFCC',
          '00FF0000', '0000FF00', '00660066']
fillers = []
 

for color in colors:
    temp = PatternFill(patternType='solid',
                       fgColor=color)
    fillers.append(temp)


cell_ids = ['B2', 'B3', 'B4', 'B5', 'A2']
for i in range(5):
    ws[cell_ids[i]].fill = fillers[i]

ws["A1"].fill=PatternFill(patternType='solid',
                       fgColor="ff22b800")

wb.save("GFGCoursePrices.xlsx")
