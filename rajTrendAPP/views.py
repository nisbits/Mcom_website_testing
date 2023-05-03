from django.shortcuts import render

# Create your views here.
from django.shortcuts import render
from mcom_website.settings import MEDIA_ROOT,MEDIA_URL
from django.core.files.storage import FileSystemStorage
from rest_framework.decorators import api_view
from datetime import date, timedelta
import datetime 
from rest_framework.response import Response
import pandas as pd
import openpyxl
import os

@api_view(['POST'])
def old_raj_trend(request):
    raw_kpi_4G=request.FILES["raw_kpi"] if "raw_kpi" in request.FILES else None
    print("raw------------------",raw_kpi_4G)
    
    location=MEDIA_ROOT+r'\trends\temporary_files'
    print('location-------------------',location)
    fs=FileSystemStorage(location=location)
    file=fs.save(raw_kpi_4G.name,raw_kpi_4G)
    file_path=fs.path(file)
    df=pd.read_excel(file_path)
    print(df)
    os.remove(path=file_path)
   

################################## for site_list ############################################
    excel_site_list_raj=request.FILES["site_list"] if "site_list" in request.FILES else None
    if excel_site_list_raj:
        location=MEDIA_ROOT+r'\trends\temporary_files'
        fs_sheet=FileSystemStorage(location=location)
        file_sheet=fs_sheet.save(excel_site_list_raj.name,excel_site_list_raj)
        file_path_site=fs_sheet.path(file_sheet)
        df_site_list=pd.read_excel(file_path_site)
        print(df_site_list)
        os.remove(path=file_path_site)

######################################################################        
    door_raj_path=os.path.join(MEDIA_ROOT,'trends','raj')
    # save_raj_path=os.path.join(door_raj_path,"actual input/RAJ_TOOL_KPI.xlsx")
    # df = pd.read_excel(save_raj_path)
   
    
    STR=os.path.join(door_raj_path,'template','update raj trend.xlsx')
    wb=openpyxl.load_workbook(STR)
    ws=wb.active

    df["Short name"] = df["Short name"].fillna(method="ffill")  #######for (forward fill)ffill is used to copy  and fill
    df.columns.values[1] = 'Date'  #####for empty column

    df["DL User Throughput_Kbps [CDBH]"] = (df["DL User Throughput_Kbps [CDBH]"]/1024)  #####for change kbps to mbps
    df.columns.values[4] = 'DL User Throughput_Mbps [CDBH]'
    # print(df)

    sn=[]
    techlist=[]
    strip=[]

    for cell in df['Short name']:
        if("_" in cell):
            sit_id = cell.split("_")[-2][:-1]
            sn.append(sit_id)
        

        else:
            sit_id=cell[:-1]
            sn.append(sit_id)
        
        if('_F1_' in cell or '_F3_' in cell or '_F8_' in cell  or '_T1' in cell or '_T2_' in cell ):  
            if('_F1_' in cell):
                tech='L2100'
                # techlist.append(tech)
            if('_F3_' in cell):
                tech='L1800'
                # techlist.append(tech) 
            if('_f8_' in cell):
                tech='L900'    
                # techlist.append(tech)
            if('_T1_' in cell or '_T2_' in cell):
                tech='L2300'
            techlist.append(tech)
                            
        else:
            cell=sit_id
            techlist.append(cell)
        
    # if ('' in cell):
    #     txt=cell.strip()
    #     strip.append(txt)
    # else:
    #     txt=cell()
    #     strip.append(txt)      
        
    # print(techlist)
    # df.insert(2,'site type',sn)
    df.insert(0, 'Site ID', sn)
    df.insert(4,'Tech',techlist)  
    # df.insert(5,'strip',strip)

    df["Short name"] =df["Short name"].apply(lambda x: x.strip())  


    df.fillna(value=0,inplace=True)  
    
    raj_1st_step=os.path.join(door_raj_path,'process output','fill.xlsx')
    df.to_excel(raj_1st_step, index=False)
    print(df)
    excel_file_1 = raj_1st_step
    # raj_2nd_step=os.path.join(door_raj_path,'project file','site.xlsx')
    # excel_file_2 = raj_2nd_step

    df1 = pd.read_excel(excel_file_1)
    # df2 = pd.read_excel(excel_file_2)

    df1.rename(columns={"Site ID": "Site_ID"}, inplace=True)
    # df_site_list.rename(columns={"Site ID": "Site_ID"}, inplace=True)

    kpi = ['Radio NW Availability', 'E-UTRAN Average CQI [CDBH]', 'DL User Throughput_Mbps [CDBH]', 'UL RSSI [CDBH]',
        'PS Drop Call Rate % [CDBH]', 'RRC Setup Success Rate [CDBH]', 'ERAB Setup Success Rate [CDBH]',
        'PS handover success rate [LTE Intra System] [CDBH]', 'MV_Average number of used DL PRBs [CDBH]',
        'MV_UL User Throughput_Kbps [CDBH]', 'MV_VoLTE Traffic', 'MV_VoLTE Packet Loss DL [CBBH]',
        'MV_VoLTE Packet Loss UL [CBBH]', 'MV_VoLTE DCR [CBBH]', 'MV_CSFB Redirection Success Rate',
        'MV_VoLTE ERAB Setup Success Rate', 'MV_4G Data Volume_GB',
        'PS handover success rate [LTE Inter System] [CDBH]', 'VoLTE InterF HOSR Exec [CBBH]',
        'VoLTE IntraF HOSR Exec [CBBH]', 'PS Drop Call Rate % [CBBH]','VoLTE DCR_Nom [CBBH]','CA Data Volume [MB]']

    filtered_df_1 = df1[(df1.Site_ID.isin(list(df_site_list['2G ID'])))]

    print(filtered_df_1)
    raj_filtered_3rd_step=os.path.join(door_raj_path,'process output','filtered_df_1.xlsx')
    filtered_df_1.to_excel(raj_filtered_3rd_step, index=False)

    df1 = pd.read_excel(raj_filtered_3rd_step)
    df_pivot = df1.pivot_table(values=kpi, columns='Date', index=['Short name', 'Site_ID','Tech'])
    
    
    raj_pivot_4th_step=os.path.join(door_raj_path,'process output','pivot.xlsx')
    df_pivot.to_excel(raj_pivot_4th_step)
    
    alpha = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    def num_hash(num):
            if num < 26:
                return alpha[num-1]
            else:
                q, r = num//26, num % 26
                if r == 0:
                    if q == 1:
                        return alpha[r-1]
                    else:
                        return num_hash(q-1) + alpha[r-1]
                else:
                    return num_hash(q) + alpha[r-1]
            
            
    def titleToNumber(s):
	# This process is similar to binary-to-
	# decimal conversion
        result = 0
        for B in range(len(s)):
            result *= 26
            result += ord(s[B]) - ord('A') + 1
        return result
    str_date=request.POST.get('offered_date')
    date1 = datetime.datetime.strptime(str_date, '%Y-%m-%d')
    print("Date-------------------------------",date1)
    # date1=cal.get_date()
    # date1=date(2023,3,2)
    d1=date1-timedelta(1)
    d2=date1-timedelta(2)
    d3=date1-timedelta(3)
    d4=date1-timedelta(4)
    d5=date1-timedelta(5)
    cl=[d1,d2,d3,d4,d5]
   
    def overwrite(kpi_name,coln1):
        coln2=num_hash(titleToNumber(coln1)+1)
        coln3=num_hash(titleToNumber(coln1)+2)
        coln4=num_hash(titleToNumber(coln1)+3)
        coln5=num_hash(titleToNumber(coln1)+4)
        print(kpi_name)
        index=df_pivot.index
        print('donnnnnnnnnnnnnnnnne')
        print(len(index))
        dr=df_pivot[kpi_name]
        
        li=dr.columns
        col1=dr[li[0]].to_list()
        col2=dr[li[1]].to_list()
        col3=dr[li[2]].to_list()
        col4=dr[li[3]].to_list()
        col5=dr[li[4]].to_list()


        ws[coln1+"4"].value=cl[4]
        ws[coln2+"4"].value=cl[3]
        ws[coln3+"4"].value=cl[2]
        ws[coln4+"4"].value=cl[1]
        ws[coln5+"4"].value=cl[0]



        for i,value in enumerate(index):
                j=i+6
                ws['N'+str(j)].value='CAPACITY'
                ws['M'+str(j)].value='ERICSSON'
                ws['K'+str(j)].value=date1

                ws['L'+str(j)].value='DONE'
                ws['B'+str(j)].value=index[i][1]
                ws['C'+str(j)].value=index[i][0]
                ws['G'+str(j)].value=index[i][0]
                ws['I'+str(j)].value=index[i][2]
                
                # ws['D'+str(j)].value=index[i][2]
                # ws['E'+str(j)].value=index[i][2]
                ws['F'+str(j)].value=index[i][0]     
                ws[coln1+str(j)].value=col1[i]
                ws[coln2+str(j)].value=col2[i]
                ws[coln3+str(j)].value=col3[i]
                ws[coln4+str(j)].value=col4[i]
                ws[coln5+str(j)].value=col5[i]
    for kpi_name in kpi:
        if(kpi_name=='RRC Setup Success Rate [CDBH]'):
            overwrite(kpi_name,'AM')

        if(kpi_name=='ERAB Setup Success Rate [CDBH]'):
            overwrite(kpi_name,'AT') 

        if(kpi_name=='PS Drop Call Rate % [CDBH]'):
            overwrite(kpi_name,'BA')

        if(kpi_name=='DL User Throughput_Mbps [CDBH]'):
            overwrite(kpi_name,'BH')

        if(kpi_name=='MV_UL User Throughput_Kbps [CDBH]'):
            overwrite(kpi_name,'BO') 
            
        if(kpi_name=='PS handover success rate [LTE Intra System] [CDBH]'):
            overwrite(kpi_name,'BV') 

        if(kpi_name=='PS handover success rate [LTE Inter System] [CDBH]'):
            overwrite(kpi_name,'CC')    

        if(kpi_name=='MV_CSFB Redirection Success Rate'):
            overwrite(kpi_name,'CJ') 

        if(kpi_name=='MV_VoLTE ERAB Setup Success Rate'):
            overwrite(kpi_name,'CQ') 

        if(kpi_name=='MV_VoLTE DCR [CBBH]'):
            overwrite(kpi_name,'CX')    

        if(kpi_name=='MV_VoLTE Packet Loss DL [CBBH]'):
            overwrite(kpi_name,'DE') 

        if(kpi_name=='MV_VoLTE Packet Loss UL [CBBH]'):
            overwrite(kpi_name,'DL')  

        if(kpi_name=='VoLTE IntraF HOSR Exec [CBBH]'):
            overwrite(kpi_name,'DZ')   

        if(kpi_name=='VoLTE InterF HOSR Exec [CBBH]'):
            overwrite(kpi_name,'EG')  

        if(kpi_name=='E-UTRAN Average CQI [CDBH]'):
            overwrite(kpi_name,'EN')  

        if(kpi_name=='UL RSSI [CDBH]'):
            overwrite(kpi_name,'EU')   

        if(kpi_name=='MV_4G Data Volume_GB'):
            overwrite(kpi_name,'FB')  

        if(kpi_name=='MV_VoLTE Traffic'):
            overwrite(kpi_name,'FK')  

        if(kpi_name=='MV_Average number of used DL PRBs [CDBH]'):
            overwrite(kpi_name,'FR')    

        if(kpi_name=='Radio NW Availability'):
            overwrite(kpi_name,'FX') 
            
        if(kpi_name=='VoLTE DCR_Nom [CBBH]'):
            overwrite(kpi_name,'GD') 
            
        if(kpi_name=='CA Data Volume [MB]'):
            overwrite(kpi_name,'GJ')         
    save_raj_trend=os.path.join(door_raj_path,'output','Rajasthantrend.xlsx')    
    print(save_raj_trend,'-----------------------------------------')
    wb.save(save_raj_trend) 
    Download_path=os.path.join(MEDIA_URL,'trends','raj','output','Rajasthantrend.xlsx')
    return Response({"message":"Succesfully uploaded",'status':True,'Download_url':Download_path})
   
       

    