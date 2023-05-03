from django.shortcuts import render
from rest_framework.decorators import api_view
from rest_framework.response import Response
from django.core.files.storage import FileSystemStorage
from datetime import date,timedelta
import datetime
import openpyxl
from mcom_website.settings import MEDIA_ROOT,MEDIA_URL
import pandas as pd
import os

@api_view(['POST'])
def old_kol_trend(request):
    raw_kpi_4G=request.FILES['raw_kpi'] if "raw_kpi" in request.FILES else None
    if raw_kpi_4G:
        location=MEDIA_ROOT+r'\trends\temporary_files'
        fs=FileSystemStorage(location=location)
        file=fs.save(raw_kpi_4G.name,raw_kpi_4G)
        file_path=fs.path(file)
        df_raw_kpi=pd.read_excel(file_path)
        print("________________df raw Kpi_______________",df_raw_kpi)
        os.remove(path=file_path)
    #     return Response({'status':True})
    # else:
    #     return Response({'status':False})

    site_list_4G=request.FILES["site_list"] if "site_list" in request.FILES else None
    if site_list_4G:
        location=MEDIA_ROOT+r'\trends\temporary_files'
        fs=FileSystemStorage(location=location)
        file=fs.save(site_list_4G.name,site_list_4G)
        file_site_path=fs.path(file)
        df_site_list=pd.read_excel(file_site_path)
        print("______________df_sit_list__________",df_site_list)
        os.remove(path=file_site_path)
#####################################################################################

    
    
    print("here")
    df_raw_kpi['Short name'].fillna(inplace=True, method="ffill")
    df_raw_kpi["Short name"] =df_raw_kpi["Short name"].apply(lambda x: x.strip()) # to remove all tralling spaces..... and leading spaces..
    
    # df_raw_kpi.columns.values[1]='Date'
    df_raw_kpi.rename( columns={'Unnamed: 1':'date'}, inplace=True )
    df_raw_kpi['MV_DL User Throughput_Kbps [CDBH]']=(df_raw_kpi['MV_DL User Throughput_Kbps [CDBH]']/1024)
    # df_raw_kpi.columns.values[6]='MV_DL User Throughput_Mbps [CDBH]'
    df_raw_kpi.rename(columns={"MV_DL User Throughput_Kbps [CDBH]" :"MV_DL User Throughput_Mbps [CDBH]" } ,inplace = True )
    
    print(df_raw_kpi)

    split=[]
    tech=[]
    for cell in df_raw_kpi['Short name']:
        if('_' in cell):
            site_id = cell.split("_")[-2][:-1]
            split.append(site_id)
            
        else:
            site_id=cell[:-1]
            split.append(site_id) 
            
        if('KO' in cell or 'WB' in cell ):  
            if('KO' in cell):
                circle='KOL' 
                tech.append(circle)
            if('WB' in cell ):
                circle='ROB'
                tech.append(circle)                   
        else:
            circle=site_id
            tech.append(circle)     
            
    df_raw_kpi.insert(0,'Site ID', split)
    df_raw_kpi.insert(6,'circle',tech )  
    
    # print(df)

    ecgi=[]
    cgi=[]
    for cell1 in df_raw_kpi['4G_ECGI']:
        if('-' in cell1):
            lnbts_id = cell1.split("-")[-2]
            ecgi.append(lnbts_id)   
        else:
            lnbts_id=cell1
            ecgi.append(lnbts_id)
            
        if('-' in cell1):
            lnbts_Name=cell1.split('-')[-1]
            cgi.append(lnbts_Name) 
        else:
            lnbts_Name=cell1
                    
    df_raw_kpi.insert(2,'ENODEB ID', ecgi)
    df_raw_kpi.insert(3,'ci',cgi) 
    print(df_raw_kpi) 
    df_raw_kpi.fillna(value=0,inplace=True)  
    door_path=os.path.join(MEDIA_ROOT,'trends','kol')

    PsOsPath1=os.path.join(door_path,'process output','desird_output.xlsx')    
    df_raw_kpi.to_excel(PsOsPath1,index=False)

    fill_excelfile_1=PsOsPath1
    
    # PsOsPath2=os.path.join(door_path,'project file','site.xlsx')
    # check_excelfile_2=PsOsPath2

    df1=pd.read_excel(fill_excelfile_1)
    # df2=pd.read_excel(check_excelfile_2,sheet_name='Sheet2')

    df1.rename(columns={"ENODEB ID": "ENODEB_ID"}, inplace=True)
    # df_site_list.rename(columns={"ENODEB ID": "ENODEB_ID"}, inplace=True)

    kpi=['RRC Setup Success Rate [CBBH]','VoLTE ERAB Setup Success Rate [CBBH]','PS Drop Call Rate % [CDBH]_Old','MV_DL User Throughput_Mbps [CDBH]',
    'MV_UL User Throughput_Kbps [CDBH]','PS handover success rate [LTE Intra System] [CDBH]','PS handover success rate [LTE Inter System] [CBBH]',
    'MV_CSFB Redirection Success Rate [CDBH]','ERAB Setup Success Rate [CDBH]','VoLTE DCR [CBBH]','VoLTE Packet Loss DL [CBBH]',
    'VoLTE Packet Loss UL [CBBH]','VoLTE Intra-LTE Handover Success Ratio [CBBH]','VoLTE Inter-Frequency Handover Success Ratio [CBBH]',
    'E-UTRAN Average CQI [CDBH]','UL RSSI [CDBH]','4G Data Volume [GB]','VoLTE Traffic','Average number of used DL PRBs','VoLTE SRVCC SR']

    filtered_df_1 = df1[(df1.ENODEB_ID.isin(list(df_site_list['2G ID'])))]

    print(filtered_df_1)
    PsOs_Filter=os.path.join(door_path,'process output','filtered.xlsx')
    filtered_df_1.to_excel(PsOs_Filter, index=False)

    df1 = pd.read_excel(PsOs_Filter)
    df_pivot = df1.pivot_table(values=kpi,columns='date', index=['Short name', 'Site ID','circle','ENODEB_ID','ci','4G_ECGI'])
    
    PsOs_Pivot=os.path.join(door_path,"process output','pivot.xlsx")
    df_pivot.to_excel(PsOs_Pivot) 
    
    alpha = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    def num_hash(num):
            if num < 26:
                return alpha[num-1]
            else:
                q, r = num//26, num % 26
                if r == 0:
                    if q == 1:
                        return alpha[r-1]
                    else:
                        return num_hash(q-1) + alpha[r-1]
                else:
                    return num_hash(q) + alpha[r-1]
            
            
    def titleToNumber(s):
	# This process is similar to binary-to-
	# decimal conversion
        result = 0
        for B in range(len(s)):
            result *= 26
            result += ord(s[B]) - ord('A') + 1
        return result
    str_date=request.POST.get("offered_date")
    date1=datetime.datetime.strptime(str_date,"%Y-%m-%d")
    print("strdate--------------------------------------",date1)

    # date1 =date(2023,1,27)
    # date1=cal.get_date()
    d1=date1-timedelta(1)
    d2=date1-timedelta(2)
    d3=date1-timedelta(3)
    d4=date1-timedelta(4)
    d5=date1-timedelta(5)
    cl=[d1,d2,d3,d4,d5]
    # index=df_pivot.index
    
    STR=os.path.join(door_path,'template','new up.xlsx')
    wb=openpyxl.load_workbook(STR)
    ws=wb.active
    
    def overwrite(kpi_name,coln1):
        coln2=num_hash(titleToNumber(coln1)+1)
        coln3=num_hash(titleToNumber(coln1)+2)
        coln4=num_hash(titleToNumber(coln1)+3)
        coln5=num_hash(titleToNumber(coln1)+4)
        print(kpi_name)
        index=df_pivot.index
        print('donnnnnnnnnnnnnnnnne')
        print(len(index))
        dr=df_pivot[kpi_name]
        li=dr.columns
        col1=dr[li[0]].to_list()
        col2=dr[li[1]].to_list()
        col3=dr[li[2]].to_list()
        col4=dr[li[3]].to_list()
        col5=dr[li[4]].to_list()

        ws[coln1+"4"].value=cl[4]
        ws[coln2+"4"].value=cl[3]
        ws[coln3+"4"].value=cl[2]
        ws[coln4+"4"].value=cl[1]
        ws[coln5+"4"].value=cl[0]

        for i,value in enumerate(index):
                j=i+5
                ws['B'+str(j)].value='ULS'
                ws['A'+str(j)].value=index[i][2]
                ws['D'+str(j)].value=index[i][1]
                ws['E'+str(j)].value=index[i][3]
                ws['F'+str(j)].value=index[i][0]
                ws['G'+str(j)].value=index[i][4]
                ws['H'+str(j)].value=index[i][5]
            
                
                
                ws[coln1+str(j)].value=col1[i]
                ws[coln2+str(j)].value=col2[i]
                ws[coln3+str(j)].value=col3[i]
                ws[coln4+str(j)].value=col4[i]
                ws[coln5+str(j)].value=col5[i]
                
    for kpi_name in kpi:
        
        if(kpi_name=='RRC Setup Success Rate [CBBH]'):
            overwrite(kpi_name,'I') 
            
        if(kpi_name=='VoLTE ERAB Setup Success Rate [CBBH]'):
            overwrite(kpi_name,'P')    
            
        if(kpi_name=='PS Drop Call Rate % [CDBH]_Old'):
            overwrite(kpi_name,'W')   
            
        if(kpi_name=='MV_DL User Throughput_Mbps [CDBH]'):
            overwrite(kpi_name,'AD')
            
        if(kpi_name=='MV_UL User Throughput_Kbps [CDBH]'):
            overwrite(kpi_name,'AK')
            
        if(kpi_name=='PS handover success rate [LTE Intra System] [CDBH]'):
            overwrite(kpi_name,'AR')
            
        if(kpi_name=='PS handover success rate [LTE Inter System] [CBBH]'):
            overwrite(kpi_name,'AY')  
            
        if(kpi_name=='MV_CSFB Redirection Success Rate [CDBH]'):
            overwrite(kpi_name,'BF')
            
        if(kpi_name=='ERAB Setup Success Rate [CDBH]'):
            overwrite(kpi_name,'BM') 
            
        if(kpi_name=='VoLTE DCR [CBBH]'):
            overwrite(kpi_name,'BT') 
            
        if(kpi_name=='VoLTE Packet Loss DL [CBBH]'):
            overwrite(kpi_name,'CA')
            
        if(kpi_name=='VoLTE Packet Loss UL [CBBH]'):
            overwrite(kpi_name,'CH')
            
        if(kpi_name=='VoLTE SRVCC SR'):
            overwrite(kpi_name,'CO')
            
        if(kpi_name=='VoLTE Intra-LTE Handover Success Ratio [CBBH]'):
            overwrite(kpi_name,'CV')
            
        if(kpi_name=='VoLTE Inter-Frequency Handover Success Ratio [CBBH]'):
            overwrite(kpi_name,'DC')
            
        if(kpi_name=='E-UTRAN Average CQI [CDBH]'):
            overwrite(kpi_name,'DJ')
            
        if(kpi_name=='UL RSSI [CDBH]'):
            overwrite(kpi_name,'DQ')
            
        if(kpi_name=='4G Data Volume [GB]'):
            overwrite(kpi_name,'DX')
            
        if(kpi_name=='VoLTE Traffic'):
            overwrite(kpi_name,'EE') 
            
        if(kpi_name=='Average number of used DL PRBs'):
            overwrite(kpi_name,'EL')
    save_output=os.path.join(door_path,'output','KOL_KPI_TREND_OUTPUT.xlsx')    
    wb.save(save_output)
    print('successfully') 
    download_path=os.path.join(MEDIA_URL,"trends",'kol','output','KOL_KPI_TREND_OUTPUT.xlsx')
    return Response({"status":True,"message":"file uploaded sucessfully",'Download_url':download_path})   

