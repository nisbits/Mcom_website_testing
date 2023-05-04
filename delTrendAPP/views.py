import pandas as pd
import datetime
from datetime import date, timedelta
from openpyxl import Workbook,load_workbook
from openpyxl.utils import  get_column_letter
from mcom_website.settings import MEDIA_ROOT,MEDIA_URL
from rest_framework.decorators import api_view
from rest_framework.response import Response
from django.core.files.storage import FileSystemStorage
# from zipfile import ZipFile
from tkinter import *
import tkinter as tk
from tkinter import filedialog,messagebox,ttk
import openpyxl as xl
                      
import os

import glob



@api_view(['POST'])
def old_del_trend(request):
    raw_kpi_4G=request.FILES['raw_kpi_4G'] if 'raw_kpi_4G' in request.FILES else None
    print("raw------------------",raw_kpi_4G)

    if raw_kpi_4G:
        location=MEDIA_ROOT +r'\trends\temporary_files'
        print('location-------------------',location)
        fs=FileSystemStorage(location=location)
        file=fs.save(raw_kpi_4G.name,raw_kpi_4G)
        file_path=fs.path(file)
        df_raw_kpi_4G=pd.read_excel(file_path)
        
        print(df_raw_kpi_4G)
        os.remove(path=file_path)
    print('_____________hi__________________')

    site_list_4G=request.FILES['site_list_4G'] if 'site_list_4G' in request.FILES else None
    if site_list_4G:
        location=MEDIA_ROOT+r'\trends\temporary_files'
        fs=FileSystemStorage(location=location)
        file_sheet=fs.save(site_list_4G.name,site_list_4G)
        file_path=fs.path(file_sheet)
        df_site_4G=pd.read_excel(file_path)
        print(df_site_4G)
        os.remove(path=file_path)

    raw_kpi_2G=request.FILES['raw_kpi_2G'] if 'raw_kpi_2G' in request.FILES else None
    print('________________________________raw__________________')
    if raw_kpi_2G:
        location=MEDIA_ROOT +r'\trends\temporary_files'
        print("_____________location",location)
        fs=FileSystemStorage(location=location)
        file_2G=fs.save(raw_kpi_2G.name,raw_kpi_2G)
        file_2G_path=fs.path(file_2G)
        df_raw_kpi_2G=pd.read_excel(file_2G_path)
        print(df_raw_kpi_2G)
        os.remove(path=file_2G_path)

    site_list_2G=request.FILES['site_list_2G'] if 'site_list_2G' in request.FILES else None
    if site_list_2G:
        location=MEDIA_ROOT+r'\trends\temporary_files'
        fs=FileSystemStorage(location=location)
        file_2G_sheet=fs.save(site_list_2G.name,site_list_2G)
        file_2G_path=fs.path(file_2G_sheet)
        df_2G_site=pd.read_excel(file_2G_path)
        print(df_2G_site)
        os.remove(path=file_2G_path)    

    
    door_path=os.path.join(MEDIA_ROOT,'trends','del','del4G')

    # str=os.path.join(door_path,'inputs','SITES.xlsx')
    # df_site=pd.read_excel(str)
   



    kpi=[   "MV_RRC Setup Success Rate",
            "ERAB Setup Success Rate [CDBH]",
            "PS Drop Call Rate % [CBBH]",
            "MV_DL User Throughput_Kbps [CDBH]",
            "MV_UL User Throughput_Kbps [CDBH]",
            "PS handover success rate [LTE Intra System] [CDBH]",
            "PS handover success rate [LTE Inter System] [CDBH]",
            "MV_4G Data Volume_GB",
            "MV_CSFB Redirection Success Rate [CDBH]",
            "RRC Paging Discard Ratio",
            "MV_Average number of used DL PRBs [CDBH]",
            "VoLTE Call Setup Success rate [CBBH]",
            "VoLTE Drop Call Rate [CBBH]]",
            "MV_VoLTE Packet Loss DL [CBBH]",
            "MV_VoLTE Packet Loss UL [CBBH]",
            "VoLTE Intra HOSR [CBBH]",
            "VoLTE InterF HOSR Exec [CBBH]",
            "VoLTE SRVCC SR Exec [CBBH]",
            "VoLTE Traffic [CDBH]",
            "MV_VoLTE Traffic",
            "4G Data Volume [GB] [CDBH]",
            "UL RSSI [CBBH]",
            "MV_E-UTRAN Average CQI [CDBH]",
        ]
        
   

    
    site_list=list(df_site_4G["2G ID"])


    df_raw_kpi_4G["Short name"].fillna( inplace=True, method="ffill")
    df_raw_kpi_4G.rename( columns={'Unnamed: 1':'date'}, inplace=True )

    # print(df_raw_kpi_4G.columns)



    lis=list(df_raw_kpi_4G["Short name"])
    sit_id_lis=[]
    cell_id_lis=[]
    for item in lis:
        if("_" in item):
            cell_id=item.split("_")[-2]
            ln=len(item.split("_")[-1])
            #print(ln)
            sit_id=item.split("_")[-2][:-ln]
        else:
            cell_id=item
            sit_id=item
        cell_id_lis.append(cell_id)
        sit_id_lis.append(sit_id)

    # print(sit_id)
    # print(cell_id_lis)

    df_raw_kpi_4G.insert(1, "SITE_ID", sit_id_lis)
    df_raw_kpi_4G.insert(2, "CELL_ID", cell_id_lis)
    
    
    # df_all_tech_kpi["Short name"] =df_all_tech_kpi["Short name"].apply(lambda x: x.strip())
    
    df_raw_kpi_4G.rename(columns={"Short name" :"Shortname" } ,inplace = True )
    # df_all_tech_kpi.rename(columns={"site id" :"site_id" } ,inplace = True )
    df_raw_kpi_4G.fillna(value=0,inplace=True)
    
    # df_all_tech_kpi = df_all_tech_kpi.sort_values(by=['SITE_ID',"Shortname",],axis=0, ascending=True)
    
    enb=[]
    for cell in df_raw_kpi_4G['SITE_ID']:
        if pd.isnull(cell):
            pass
        else:
            enbid=cell[1:]
            enb.append(enbid)
    df_raw_kpi_4G.insert(3,'ENBID',enb)
    
    add=[]
    for i in df_raw_kpi_4G['SITE_ID']:
        ad=9
        add.append(ad)
    df_raw_kpi_4G.insert(4,'add',add)
    
    df_raw_kpi_4G['SECTOR']=df_raw_kpi_4G['add'].astype(str)  + df_raw_kpi_4G['ENBID']
    df_raw_kpi_4G
    
    PsOs_path1=os.path.join(door_path,'process_outputs','desired_input.xlsx')
    df_raw_kpi_4G.to_excel(PsOs_path1)

    str_date=request.POST.get("offered_date")
    date1=datetime.datetime.strptime(str_date,'%Y-%m-%d')
    dt1 = date1 - timedelta(1)
    dt2 = date1 - timedelta(2)
    dt3 = date1 - timedelta(3)
    dt4 = date1 - timedelta(4)
    dt5 = date1 - timedelta(5)
    ls=[dt1,dt2,dt3,dt4,dt5]

    def perticular_tech(tech,site_list):
        df_filtered = df_raw_kpi_4G[(df_raw_kpi_4G.SITE_ID.isin(site_list)) & (df_raw_kpi_4G.date.isin(ls))& (df_raw_kpi_4G.Shortname.str.contains('|'.join(tech)))]
        # print(df_filtered)
        PsOs_filter=os.path.join(door_path,'process_outputs','last_filtered_input.xlsx')
        df_filtered.to_excel(PsOs_filter)
        df_pivoted = df_filtered.pivot_table(index=["SITE_ID","Shortname","CELL_ID",'ENBID','SECTOR'], columns="date")
        # print(df_pivoted)
        save_name=str(tech) +"_pivot.xlsx"
        print("_________________________savename_________________________",tech)
        PsOs_pivot=os.path.join(door_path,'process_outputs',save_name)
       
        df_pivoted.to_excel(PsOs_pivot)
        return df_filtered,df_pivoted


    alpha = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    def num_hash(num):
            if num < 26:
                return alpha[num-1]
            else:
                q, r = num//26, num % 26
                if r == 0:
                    if q == 1:
                        return alpha[r-1]
                    else:
                        return num_hash(q-1) + alpha[r-1]
                else:
                    return num_hash(q) + alpha[r-1]
 
    def titleToNumber(s):

        result = 0
        for B in range(len(s)):
            result *= 26
            result += ord(s[B]) - ord('A') + 1
        return result


    def overwrite(df_pivoted,kpi_name,coln1,trend_ws):
        coln2=num_hash(titleToNumber(coln1)+1)
        coln3=num_hash(titleToNumber(coln1)+2)
        coln4=num_hash(titleToNumber(coln1)+3)
        coln5=num_hash(titleToNumber(coln1)+4)
        print(kpi_name)
        index_pivot=df_pivoted.index.to_list()
        print("index ;###############################",index_pivot)
        print(len(index_pivot))
        print("index of pivoted table: ",index_pivot)
        dr=df_pivoted[kpi_name]
        print("columns of dr table",dr.columns)
        cl=dr.columns.to_list()
        print("column list",cl)
        
        # site_id=dr["SITE_ID"].to_list() 
        # cell_id=dr["CELL_ID"].to_list()    
        col1=dr[str(cl[0])].to_list()
        col2=dr[str(cl[1])].to_list()
        col3=dr[str(cl[2])].to_list()
        col4=dr[str(cl[3])].to_list()
        col5=dr[str(cl[4])].to_list()

        trend_ws[coln1+"2"].value=cl[0]
        trend_ws[coln2+"2"].value=cl[1]
        trend_ws[coln3+"2"].value=cl[2]
        trend_ws[coln4+"2"].value=cl[3]
        trend_ws[coln5+"2"].value=cl[4]

        # me=column_index_from_string(coln5)+1
        # me=get_column_letter(me)
        for i,value in enumerate(index_pivot):
            j=i+3
            trend_ws["B"+str(j)].value=index_pivot[i][1]
            trend_ws["E"+str(j)].value=index_pivot[i][1]
            trend_ws["A"+str(j)].value=index_pivot[i][0] 
            trend_ws["C"+str(j)].value=index_pivot[i][3]
            trend_ws["D"+str(j)].value=index_pivot[i][4]
            trend_ws["J"+str(j)].value=date1
            trend_ws["F"+str(j)].value='DL'
            
            
            trend_ws[coln1+str(j)].value=col1[i]
            trend_ws[coln2+str(j)].value=col2[i]
            trend_ws[coln3+str(j)].value=col3[i]
            trend_ws[coln4+str(j)].value=col4[i]
            trend_ws[coln5+str(j)].value=col5[i]
            # trend_ws[me+str(j)].value='=COUNTIF(P5:T5,">=98.5")'

    # for fdd
    pivot_fdd=perticular_tech(["_F3_"],site_list)[1]
    PsOs_blnk_temp=os.path.join(door_path,'template','DEL KPIs Submission_L1800.xlsx')
    # path_of_blnk_temp=PsOs_blnk_temp
    trend_wb_L1800=load_workbook(PsOs_blnk_temp)
    trend_ws=trend_wb_L1800["KPI"]
    for kpi_name in kpi:
        if(kpi_name=="MV_RRC Setup Success Rate"):
            overwrite(pivot_fdd,kpi_name,"N",trend_ws)

        if(kpi_name=="ERAB Setup Success Rate [CDBH]"):
            overwrite(pivot_fdd,kpi_name,"T",trend_ws)

        if(kpi_name=="PS Drop Call Rate % [CBBH]"):
            overwrite(pivot_fdd,kpi_name,"Z",trend_ws)

        if(kpi_name=="MV_DL User Throughput_Kbps [CDBH]"):
            overwrite(pivot_fdd,kpi_name,"AF",trend_ws)

        if(kpi_name=="MV_UL User Throughput_Kbps [CDBH]"):
            overwrite(pivot_fdd,kpi_name,"AL",trend_ws)

        if(kpi_name=="PS handover success rate [LTE Intra System] [CDBH]"):
            overwrite(pivot_fdd,kpi_name,"AQ",trend_ws)
        
        if(kpi_name=="PS handover success rate [LTE Inter System] [CDBH]"):
            overwrite(pivot_fdd,kpi_name,"AW",trend_ws)
            
        if(kpi_name=="MV_4G Data Volume_GB"):
            overwrite(pivot_fdd,kpi_name,"DW",trend_ws)    
            
        if(kpi_name=="MV_CSFB Redirection Success Rate [CDBH]"):
            overwrite(pivot_fdd,kpi_name,"BH",trend_ws)
        
        if(kpi_name=="RRC Paging Discard Ratio"):
            overwrite(pivot_fdd,kpi_name,"BN",trend_ws)    
        
        if(kpi_name=="MV_Average number of used DL PRBs [CDBH]"):
            overwrite(pivot_fdd,kpi_name,"BS",trend_ws) 
           
        if(kpi_name=="VoLTE Call Setup Success rate [CBBH]"):
            overwrite(pivot_fdd,kpi_name,"BX",trend_ws)        
            
        if(kpi_name=="VoLTE Drop Call Rate [CBBH]]"):
            overwrite(pivot_fdd,kpi_name,"CD",trend_ws)
            
        if(kpi_name=="MV_VoLTE Packet Loss DL [CBBH]"):
            overwrite(pivot_fdd,kpi_name,"CJ",trend_ws)
                
        if(kpi_name=="MV_VoLTE Packet Loss UL [CBBH]"):
            overwrite(pivot_fdd,kpi_name,"CP",trend_ws)
            
        if(kpi_name=="VoLTE Intra HOSR [CBBH]"):
            overwrite(pivot_fdd,kpi_name,"CV",trend_ws) 
            
        if(kpi_name=="VoLTE InterF HOSR Exec [CBBH]"):
            overwrite(pivot_fdd,kpi_name,"DB",trend_ws)
            
            
        if(kpi_name=="VoLTE SRVCC SR Exec [CBBH]"):
            overwrite(pivot_fdd,kpi_name,"DH",trend_ws) 
            
        if(kpi_name=="VoLTE Traffic [CDBH]"):
            overwrite(pivot_fdd,kpi_name,"DM",trend_ws)     
        
        if(kpi_name=="MV_VoLTE Traffic"):
            overwrite(pivot_fdd,kpi_name,"DR",trend_ws)     
            
        if(kpi_name=="4G Data Volume [GB] [CDBH]"):
            overwrite(pivot_fdd,kpi_name,"BC",trend_ws)              
                
        if(kpi_name=="UL RSSI [CBBH]"):
            overwrite(pivot_fdd,kpi_name,"EB",trend_ws)    
        
        if(kpi_name=="MV_E-UTRAN Average CQI [CDBH]"):
            overwrite(pivot_fdd,kpi_name,"EG",trend_ws)
        
    # for tdd
    pivot_fdd=perticular_tech(["_T1_","_T2_"],site_list)[1]
    PsOs_blnk_temp=os.path.join(door_path,'template','DEL KPIs Submission_L2300.xlsx')  
    path_of_blnk_temp=PsOs_blnk_temp
    trend_wb_L2300=load_workbook(path_of_blnk_temp)
    # print(trend_wb.sheetnames)
    trend_ws=trend_wb_L2300["KPI"]
    for kpi_name in kpi:
        if(kpi_name=="MV_RRC Setup Success Rate"):
            overwrite(pivot_fdd,kpi_name,"N",trend_ws)

        if(kpi_name=="ERAB Setup Success Rate [CDBH]"):
            overwrite(pivot_fdd,kpi_name,"T",trend_ws)

        if(kpi_name=="PS Drop Call Rate % [CBBH]"):
            overwrite(pivot_fdd,kpi_name,"Z",trend_ws)

        if(kpi_name=="MV_DL User Throughput_Kbps [CDBH]"):
            overwrite(pivot_fdd,kpi_name,"AF",trend_ws)

        if(kpi_name=="MV_UL User Throughput_Kbps [CDBH]"):
            overwrite(pivot_fdd,kpi_name,"AL",trend_ws)

        if(kpi_name=="PS handover success rate [LTE Intra System] [CDBH]"):
            overwrite(pivot_fdd,kpi_name,"AQ",trend_ws)
        
        if(kpi_name=="PS handover success rate [LTE Inter System] [CDBH]"):
            overwrite(pivot_fdd,kpi_name,"AW",trend_ws)
            
        if(kpi_name=="MV_4G Data Volume_GB"):
            overwrite(pivot_fdd,kpi_name,"DW",trend_ws)    
            
        if(kpi_name=="MV_CSFB Redirection Success Rate [CDBH]"):
            overwrite(pivot_fdd,kpi_name,"BH",trend_ws)
        
        if(kpi_name=="RRC Paging Discard Ratio"):
            overwrite(pivot_fdd,kpi_name,"BN",trend_ws)    
        
        if(kpi_name=="MV_Average number of used DL PRBs [CDBH]"):
            overwrite(pivot_fdd,kpi_name,"BS",trend_ws) 
           
        if(kpi_name=="VoLTE Call Setup Success rate [CBBH]"):
            overwrite(pivot_fdd,kpi_name,"BX",trend_ws)        
            
        if(kpi_name=="VoLTE Drop Call Rate [CBBH]]"):
            overwrite(pivot_fdd,kpi_name,"CD",trend_ws)
            
        if(kpi_name=="MV_VoLTE Packet Loss DL [CBBH]"):
            overwrite(pivot_fdd,kpi_name,"CJ",trend_ws)
                
        if(kpi_name=="MV_VoLTE Packet Loss UL [CBBH]"):
            overwrite(pivot_fdd,kpi_name,"CP",trend_ws)
            
        if(kpi_name=="VoLTE Intra HOSR [CBBH]"):
            overwrite(pivot_fdd,kpi_name,"CV",trend_ws) 
            
        if(kpi_name=="VoLTE InterF HOSR Exec [CBBH]"):
            overwrite(pivot_fdd,kpi_name,"DB",trend_ws)
            
            
        if(kpi_name=="VoLTE SRVCC SR Exec [CBBH]"):
            overwrite(pivot_fdd,kpi_name,"DH",trend_ws) 
            
        if(kpi_name=="VoLTE Traffic [CDBH]"):
            overwrite(pivot_fdd,kpi_name,"DM",trend_ws)     
        
        if(kpi_name=="MV_VoLTE Traffic"):
            overwrite(pivot_fdd,kpi_name,"DR",trend_ws)     
            
        if(kpi_name=="4G Data Volume [GB] [CDBH]"):
            overwrite(pivot_fdd,kpi_name,"BC",trend_ws)              
                
        if(kpi_name=="UL RSSI [CBBH]"):
            overwrite(pivot_fdd,kpi_name,"EB",trend_ws)    
        
        if(kpi_name=="MV_E-UTRAN Average CQI [CDBH]"):
            overwrite(pivot_fdd,kpi_name,"EG",trend_ws)


    pivot_fdd=perticular_tech(["_F8_"],site_list)[1]
    PsOs_blnk_temp=os.path.join(door_path,'template','DEL KPIs Submission_L900.xlsx')
    path_of_blnk_temp=PsOs_blnk_temp
    trend_wb_L900=load_workbook(path_of_blnk_temp)
    # print(trend_wb.sheetnames)
    trend_ws=trend_wb_L900["KPI"]
    for kpi_name in kpi:
        if(kpi_name=="MV_RRC Setup Success Rate"):
            overwrite(pivot_fdd,kpi_name,"N",trend_ws)

        if(kpi_name=="ERAB Setup Success Rate [CDBH]"):
            overwrite(pivot_fdd,kpi_name,"T",trend_ws)

        if(kpi_name=="PS Drop Call Rate % [CBBH]"):
            overwrite(pivot_fdd,kpi_name,"Z",trend_ws)

        if(kpi_name=="MV_DL User Throughput_Kbps [CDBH]"):
            overwrite(pivot_fdd,kpi_name,"AF",trend_ws)

        if(kpi_name=="MV_UL User Throughput_Kbps [CDBH]"):
            overwrite(pivot_fdd,kpi_name,"AL",trend_ws)

        if(kpi_name=="PS handover success rate [LTE Intra System] [CDBH]"):
            overwrite(pivot_fdd,kpi_name,"AQ",trend_ws)
        
        if(kpi_name=="PS handover success rate [LTE Inter System] [CDBH]"):
            overwrite(pivot_fdd,kpi_name,"AW",trend_ws)
            
        if(kpi_name=="MV_4G Data Volume_GB"):
            overwrite(pivot_fdd,kpi_name,"DW",trend_ws)    
            
        if(kpi_name=="MV_CSFB Redirection Success Rate [CDBH]"):
            overwrite(pivot_fdd,kpi_name,"BH",trend_ws)
        
        if(kpi_name=="RRC Paging Discard Ratio"):
            overwrite(pivot_fdd,kpi_name,"BN",trend_ws)    
        
        if(kpi_name=="MV_Average number of used DL PRBs [CDBH]"):
            overwrite(pivot_fdd,kpi_name,"BS",trend_ws) 
           
        if(kpi_name=="VoLTE Call Setup Success rate [CBBH]"):
            overwrite(pivot_fdd,kpi_name,"BX",trend_ws)        
            
        if(kpi_name=="VoLTE Drop Call Rate [CBBH]]"):
            overwrite(pivot_fdd,kpi_name,"CD",trend_ws)
            
        if(kpi_name=="MV_VoLTE Packet Loss DL [CBBH]"):
            overwrite(pivot_fdd,kpi_name,"CJ",trend_ws)
                
        if(kpi_name=="MV_VoLTE Packet Loss UL [CBBH]"):
            overwrite(pivot_fdd,kpi_name,"CP",trend_ws)
            
        if(kpi_name=="VoLTE Intra HOSR [CBBH]"):
            overwrite(pivot_fdd,kpi_name,"CV",trend_ws) 
            
        if(kpi_name=="VoLTE InterF HOSR Exec [CBBH]"):
            overwrite(pivot_fdd,kpi_name,"DB",trend_ws)
            
            
        if(kpi_name=="VoLTE SRVCC SR Exec [CBBH]"):
            overwrite(pivot_fdd,kpi_name,"DH",trend_ws) 
            
        if(kpi_name=="VoLTE Traffic [CDBH]"):
            overwrite(pivot_fdd,kpi_name,"DM",trend_ws)     
        
        if(kpi_name=="MV_VoLTE Traffic"):
            overwrite(pivot_fdd,kpi_name,"DR",trend_ws)     
            
        if(kpi_name=="4G Data Volume [GB] [CDBH]"):
            overwrite(pivot_fdd,kpi_name,"BC",trend_ws)              
                
        if(kpi_name=="UL RSSI [CBBH]"):
            overwrite(pivot_fdd,kpi_name,"EB",trend_ws)    
        
        if(kpi_name=="MV_E-UTRAN Average CQI [CDBH]"):
            overwrite(pivot_fdd,kpi_name,"EG",trend_ws)

    pivot_fdd=perticular_tech(["_F1_"],site_list)[1]
    PsOs_blnk_temp=os.path.join(door_path,'template','DEL KPIs Submission_L2100.xlsx')
    path_of_blnk_temp=PsOs_blnk_temp
    trend_wb_L2100=load_workbook(path_of_blnk_temp)
    # print(trend_wb.sheetnames)
    trend_ws=trend_wb_L2100["KPI"]
    for kpi_name in kpi:
       if(kpi_name=="MV_RRC Setup Success Rate"):
            overwrite(pivot_fdd,kpi_name,"N",trend_ws)

       if(kpi_name=="ERAB Setup Success Rate [CDBH]"):
            overwrite(pivot_fdd,kpi_name,"T",trend_ws)

       if(kpi_name=="PS Drop Call Rate % [CBBH]"):
            overwrite(pivot_fdd,kpi_name,"Z",trend_ws)

       if(kpi_name=="MV_DL User Throughput_Kbps [CDBH]"):
            overwrite(pivot_fdd,kpi_name,"AF",trend_ws)

       if(kpi_name=="MV_UL User Throughput_Kbps [CDBH]"):
            overwrite(pivot_fdd,kpi_name,"AL",trend_ws)

       if(kpi_name=="PS handover success rate [LTE Intra System] [CDBH]"):
            overwrite(pivot_fdd,kpi_name,"AQ",trend_ws)
        
       if(kpi_name=="PS handover success rate [LTE Inter System] [CDBH]"):
            overwrite(pivot_fdd,kpi_name,"AW",trend_ws)
            
       if(kpi_name=="MV_4G Data Volume_GB"):
            overwrite(pivot_fdd,kpi_name,"DW",trend_ws)    
            
       if(kpi_name=="MV_CSFB Redirection Success Rate [CDBH]"):
            overwrite(pivot_fdd,kpi_name,"BH",trend_ws)
        
       if(kpi_name=="RRC Paging Discard Ratio"):
            overwrite(pivot_fdd,kpi_name,"BN",trend_ws)    
        
       if(kpi_name=="MV_Average number of used DL PRBs [CDBH]"):
            overwrite(pivot_fdd,kpi_name,"BS",trend_ws) 
            
       if(kpi_name=="VoLTE Call Setup Success rate [CBBH]"):
            overwrite(pivot_fdd,kpi_name,"BX",trend_ws)        
            
       if(kpi_name=="VoLTE Drop Call Rate [CBBH]]"):
            overwrite(pivot_fdd,kpi_name,"CD",trend_ws)
            
       if(kpi_name=="MV_VoLTE Packet Loss DL [CBBH]"):
            overwrite(pivot_fdd,kpi_name,"CJ",trend_ws)
                    
       if(kpi_name=="MV_VoLTE Packet Loss UL [CBBH]"):
            overwrite(pivot_fdd,kpi_name,"CP",trend_ws)
            
       if(kpi_name=="VoLTE Intra HOSR [CBBH]"):
            overwrite(pivot_fdd,kpi_name,"CV",trend_ws) 
            
       if(kpi_name=="VoLTE InterF HOSR Exec [CBBH]"):
            overwrite(pivot_fdd,kpi_name,"DB",trend_ws)
                
       if(kpi_name=="VoLTE SRVCC SR Exec [CBBH]"):
            overwrite(pivot_fdd,kpi_name,"DH",trend_ws) 
            
       if(kpi_name=="VoLTE Traffic [CDBH]"):
            overwrite(pivot_fdd,kpi_name,"DM",trend_ws)     
        
       if(kpi_name=="MV_VoLTE Traffic"):
            overwrite(pivot_fdd,kpi_name,"DR",trend_ws)     
            
       if(kpi_name=="4G Data Volume [GB] [CDBH]"):
            overwrite(pivot_fdd,kpi_name,"BC",trend_ws)              
                
       if(kpi_name=="UL RSSI [CBBH]"):
            overwrite(pivot_fdd,kpi_name,"EB",trend_ws)    
        
       if(kpi_name=="MV_E-UTRAN Average CQI [CDBH]"):
            overwrite(pivot_fdd,kpi_name,"EG",trend_ws)

    save_wb_L2100=os.path.join(MEDIA_ROOT,'trends','del','output','output','out_DEL KPIs Submission_L2100.xlsx')
    trend_wb_L2100.save(save_wb_L2100)
    save_wb_L2300=os.path.join(MEDIA_ROOT,'trends','del','output','output','out_DEL KPIs Submission_L2300.xlsx')
    trend_wb_L2300.save(save_wb_L2300)
    save_wb_L1800=os.path.join(MEDIA_ROOT,'trends','del','output','output','out_DEL KPIs Submission_L1800.xlsx')
    trend_wb_L1800.save(save_wb_L1800)
    save_wb_L900=os.path.join(MEDIA_ROOT,'trends','del','output','output','out_DEL KPIs Submission_L900.xlsx')
    trend_wb_L900.save(save_wb_L900)  

   
   
##############################2G##################

   

    door_path=os.path.join(MEDIA_ROOT,'trends','del','del2G')

    df_raw_kpi_2G['Short name']=df_raw_kpi_2G['Short name'].fillna(method=('ffill'))
    df_raw_kpi_2G.columns.values[1]='DATE'
    a=[]
    for cell in df_raw_kpi_2G['Short name']:
        Site_ID=cell[:-1]
        a.append(Site_ID)
    df_raw_kpi_2G.insert(3,'Site_ID',a) 
    df_raw_kpi_2G.fillna(value=0,inplace=True)                     
    df_raw_kpi_2G.rename(columns={"Short name": "Shortname"}, inplace=True)

    PsOs_gsm=os.path.join(door_path,'process_outputs','fill_2g.xlsx')
    
    df_raw_kpi_2G.to_excel(PsOs_gsm,index=False) 
    
    excel_1=PsOs_gsm
    # excel_2='inputs/2G_sites.xlsx'  
    
    df1=pd.read_excel(excel_1)
    # df2=pd.read_excel(excel_2)

    # df2.rename(columns={"2G ID": "2G_ID"}, inplace=True)
    
    gsm=[
        "Total Voice Traffic [BBH]",
        "SDCCH Drop Call Rate [BBH]",
        "SDCCH Drop Call Rate_Nom [BBH]",
        "SDCCH Blocking Rate [BBH]",
        "TCH Blocking Rate [BBH]",
        "Drop Call Rate [BBH]",
        "RX Quality [BBH]",
        "Handover Success Rate [BBH]",
        "Handover Success Rate_Nom [BBH]",
        "Handover Success Rate_Denom [BBH]",
        "TCH Assignment Success Rate [BBH]",
        "ICM Band4-5 [BBH]",
        "Number of Available TCH [BBH]",
        "TNDROP [BBH]",
    ]
    
    G2_filter=df1[(df1.Site_ID.isin(list(df_2G_site['2G ID'])))]
    print(G2_filter) 
    G2Os_filter=os.path.join(door_path,'process_outputs','2Gfilter.xlsx')
    G2_filter.to_excel(G2Os_filter,index=False)  
    
    df1=pd.read_excel(G2Os_filter)  
    G2_pivot=df1.pivot_table(values=gsm,columns='DATE',index=['Shortname','Site_ID']) 
    G2Os_pivot=os.path.join(door_path,'process_outputs','G2_pivot.xlsx')
    G2_pivot.to_excel(G2Os_pivot)
    
    str_temp=os.path.join(door_path,'template','GSM KPI-OLD.xlsx')
    STR=str_temp
    wb=xl.load_workbook(STR)
    ws1=wb.active
    
    str_date=request.POST.get('offered_date')
    date1=datetime.datetime.strptime(str_date,'%Y-%m-%d')
    d1=date1-timedelta(1)
    d2=date1-timedelta(2)
    d3=date1-timedelta(3)
    d4=date1-timedelta(4)
    d5=date1-timedelta(5)
    cl=[d1,d2,d3,d4,d5]
    
    alpha = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    def num_hash(num):
            if num < 26:
                return alpha[num-1]
            else:
                q, r = num//26, num % 26
                if r == 0:
                    if q == 1:
                        return alpha[r-1]
                    else:
                        return num_hash(q-1) + alpha[r-1]
                else:
                    return num_hash(q) + alpha[r-1] 
    def titleToNumber(s):
	# This process is similar to binary-to-
	# decimal conversion
        result = 0
        for B in range(len(s)):
            result *= 26
            result += ord(s[B]) - ord('A') + 1
        return result             
    
    
    def overwrite(gsm_name,coln1,ws):
        coln2=num_hash(titleToNumber(coln1)+1)
        coln3=num_hash(titleToNumber(coln1)+2)
        coln4=num_hash(titleToNumber(coln1)+3)
        coln5=num_hash(titleToNumber(coln1)+4)
        
        index_GSM=G2_pivot.index  
        dr=G2_pivot[gsm_name]
        li=dr.columns.to_list()
        
        col1=dr[li[0]].to_list()
        col2=dr[li[1]].to_list()
        col3=dr[li[2]].to_list()
        col4=dr[li[3]].to_list()
        col5=dr[li[4]].to_list()


        ws[coln1+"2"].value=cl[4]
        ws[coln2+"2"].value=cl[3]
        ws[coln3+"2"].value=cl[2]
        ws[coln4+"2"].value=cl[1]
        ws[coln5+"2"].value=cl[0]
        
      
        
        for i,value in enumerate(index_GSM):
                j=i+3
                ws['A'+str(j)].value='NT'
   
                ws['F'+str(j)].value='DL'
              
                # ws['E'+str(j)].value=index_GSM[i][2]
                ws['C'+str(j)].value=index_GSM[i][0]
                ws['D'+str(j)].value=index_GSM[i][1]
                
                ws[coln1+str(j)].value=col1[i]
                ws[coln2+str(j)].value=col2[i]
                ws[coln3+str(j)].value=col3[i]
                ws[coln4+str(j)].value=col4[i]
                ws[coln5+str(j)].value=col5[i]
    for gsm_name in gsm:
        if(gsm_name=='Total Voice Traffic [BBH]'):
            overwrite(gsm_name,'G',ws1)
        if(gsm_name=='SDCCH Drop Call Rate [BBH]'):
            overwrite(gsm_name,'L',ws1)  
        if(gsm_name=='SDCCH Drop Call Rate_Nom [BBH]'):
            overwrite(gsm_name,'R',ws1)  
        if(gsm_name=='SDCCH Blocking Rate [BBH]'):
            overwrite(gsm_name,'W',ws1)  
        if(gsm_name=='TCH Blocking Rate [BBH]'):
            overwrite(gsm_name,'AC',ws1)  
        if(gsm_name=='Drop Call Rate [BBH]'):
            overwrite(gsm_name,'AI',ws1)  
        if(gsm_name=='RX Quality [BBH]'):
            overwrite(gsm_name,'AO',ws1)  
        if(gsm_name=='Handover Success Rate [BBH]'):
            overwrite(gsm_name,'AU',ws1)  
        if(gsm_name=='Handover Success Rate_Nom [BBH]'):
            overwrite(gsm_name,'BA',ws1)  
        if(gsm_name=='Handover Success Rate_Denom [BBH]'):
            overwrite(gsm_name,'BF',ws1)  
        if(gsm_name=='TCH Assignment Success Rate [BBH]'):
            overwrite(gsm_name,'BK',ws1)  
        if(gsm_name=='ICM Band4-5 [BBH]'):
            overwrite(gsm_name,'BQ',ws1)  
        if(gsm_name=='Number of Available TCH [BBH]'):
            overwrite(gsm_name,'BV',ws1)  
        if(gsm_name=='TNDROP [BBH]'):
            overwrite(gsm_name,'CA',ws1) 
    save_outputs=os.path.join(MEDIA_ROOT,'del','output','output','2G_Output_Trend.xlsx')        
    wb.save(save_outputs) 
    # save_wb_L2100=os.path.join(door_path,'output','out_DEL KPIs Submission_L2100.xlsx')
    # trend_wb_L2100.save(save_wb_L2100)
    # save_wb_L2300=os.path.join(door_path,'output','out_DEL KPIs Submission_L2300.xlsx')
    # trend_wb_L2300.save(save_wb_L2300)
    # save_wb_L1800=os.path.join(door_path,'output','out_DEL KPIs Submission_L1800.xlsx')
    # trend_wb_L1800.save(save_wb_L1800)
    # save_wb_L900=os.path.join(door_path,'output','out_DEL KPIs Submission_L900.xlsx')
    # trend_wb_L900.save(save_wb_L900)


    # file_paths=[]
    # for root,directories,files in os.walk(r'/mcom_website/media/trends/del/output/output'):
    #     for filename in files:
    #         filepath=os.path.join(root,filename)
    #         file_paths.append(filepath)
    # with ZipFile(r'/mcom_website/media/trends/del/output/output/output.zip','w') as zip:
    #     for file in file_paths:
    #         zip.write(file)
     
    # download_path=os.path.join(MEDIA_URL,'trends','del','output','output','output.zip')
    download_path1=os.path.join(MEDIA_URL,'trends','del','output','output','2G_Output_Trend.xlsx')
    download_path2=os.path.join(MEDIA_URL,'trends','del','output','output','out_DEL KPIs Submission_L2100.xlsx')
    download_path3=os.path.join(MEDIA_URL,'trends','del','output','output','out_DEL KPIs Submission_L1800.xlsx')
    download_path4=os.path.join(MEDIA_URL,'trends','del','output','output','out_DEL KPIs Submission_L2300.xlsx')
    download_path5=os.path.join(MEDIA_URL,'trends','del','output','output','out_DEL KPIs Submission_L900.xlsx')






    
    

    return Response({'status':True,'message':'successfully','Download_url1':download_path1,'Download_url2':download_path2,
                     'Download_url3':download_path3,'Download_url4':download_path4,'Download_url5':download_path5})
