from django.shortcuts import render
from django.http import HttpResponse
from mcom_website.settings import MEDIA_ROOT,MEDIA_URL
from rest_framework.decorators import api_view
from rest_framework.response import Response
from tkinter import *
from django.core.files.storage import FileSystemStorage
from tkinter import filedialog, messagebox, ttk
import openpyxl
import pandas as pd
import numpy as np
from datetime import date, timedelta
import datetime

import os

from zipfile import ZipFile
import os

@api_view(["POST"])
def old_ap4G_trend(request):
    raw_kpi_4G=request.FILES['raw_kpi_4G'] if "raw_kpi_4G" in request.FILES else None
    if raw_kpi_4G:
        location=MEDIA_ROOT+r'\trends\temporary_files'
        fs=FileSystemStorage(location=location)
        file=fs.save(raw_kpi_4G.name,raw_kpi_4G)
        file_path=fs.path(file)
        df_raw_kpi_4G=pd.read_excel(file_path)
        print(df_raw_kpi_4G)
        os.remove(path=file_path)

    site_list_4G=request.FILES["site_list_4G"] if "site_list_4G" in request.FILES else None
    if site_list_4G:
        location=MEDIA_ROOT+r'\trends\temporary_files'
        fs=FileSystemStorage(location=location)
        file=fs.save(site_list_4G.name,site_list_4G)
        file_site_path=fs.path(file)
        df_site_list_4G=pd.read_excel(file_site_path)
        print(df_site_list_4G)
        os.remove(path=file_site_path)
 

    door_path=os.path.join(MEDIA_ROOT,'trends','ap','ap4G')
    # PsOsPath=os.path.join(door_path,'input','SITES.xlsx')
  
    # df_site=pd.read_excel(PsOsPath)
    g2_site=[]
    f8_site=[]
    f3_site=[]
    t1t2_site=[]


    kpii=[
    "TA Sample<500 M_Nom",
    "TA Sample<1 KM_Nom",
    "TA Sample >1.5 KM_Nom_Ericsson_1",
    "TA Sample >3.4 KM_Nom",
    "TA Sample >4.45 KM_Nom",
    "TA Sample >5.7 KM_Nom",
    "TA Sample >7.2 KM_Nom",
     "TA Sample >8.4 KM_Nom"]



    kpi=[   
    'RRC Setup Success Rate [CDBH]', 
    'ERAB Setup Success Rate [CDBH]', 
    'PS Drop Call Rate % [CDBH]',
    'MV_DL User Throughput_Mbps [CDBH]',
    'MV_Average number of used DL PRBs [CDBH]', 
    'MV_UL User Throughput_Mbps [CDBH]',
    'MV_PS handover success rate [LTE Intra System] [CDBH]', 
    'PS handover success rate [LTE Inter System] [CDBH]',
    'MV_LTE_AVG_CQI', 
    'VoLTE Call Setup Success rate [CBBH]', 
    'MV_VoLTE DCR [CBBH]', 'VoLTE IntraF HOSR Exec [CBBH]',
    'VoLTE InterF HOSR Exec [CBBH]',
    'MV_VoLTE Packet Loss DL [CBBH]',
    'MV_VoLTE Packet Loss UL [CBBH]',
    'Radio NW Availability','UL RSSI [CDBH]']


    g2_kpi=[

    "SDCCH Drop Call Rate",	
    "SDCCH Blocking Rate",	
    "TCH Blocking Rate",
    "Drop Call Rate",	
    "RX Quality",	
    "Handover Success Rate",	
    "Handover Success Rate_Denom",	
    "Handover Success Rate_Nom",	
    "TCH Assignment Success Rate",
    "Network availability [RNA]",
    "Total Voice Traffic",]

    # site_list=list(df_site["2G ID"])
    site_list=list(df_site_list_4G["2G ID"])

    save_path=os.path.join(door_path,'input','ap_tool2_kpi.xlsx')
    # df_tech=pd.read_excel("input/ap_tool2_kpi.xlsx")
    df_raw_kpi_4G=pd.read_excel(save_path)
    df_raw_kpi_4G["Short name"].fillna( inplace=True, method="ffill")
    df_raw_kpi_4G.rename( columns={'Unnamed: 1':'date'}, inplace=True )
   
    df_raw_kpi_4G["MV_DL User Throughput_Kbps [CDBH]"] =(df_raw_kpi_4G["MV_DL User Throughput_Kbps [CDBH]"]/1024)
    df_raw_kpi_4G.rename(columns={"MV_DL User Throughput_Kbps [CDBH]" :"MV_DL User Throughput_Mbps [CDBH]" } ,inplace = True )

    df_raw_kpi_4G["MV_UL User Throughput_Kbps [CDBH]"] =(df_raw_kpi_4G["MV_UL User Throughput_Kbps [CDBH]"]/1024)
    df_raw_kpi_4G.rename(columns={"MV_UL User Throughput_Kbps [CDBH]" :"MV_UL User Throughput_Mbps [CDBH]"} ,inplace = True )

    # print(df_tech)
    # print(df_tech.columns)

    lis=list(df_raw_kpi_4G["Short name"])
    sit_id_lis=[]
    cell_id_lis=[]
    for item in lis:
        if("_" in item):
            cell_id=item.split(" ")[0]
            ln=len(item.split("_")[-1])
            #print(ln)
            sit_id=item.split("_")[-2][:-ln]
        else:
            cell_id=item
            sit_id=item
        cell_id_lis.append(cell_id)
        sit_id_lis.append(sit_id)

    # print(sit_id)
    # print(cell_id_lis)

    df_raw_kpi_4G.insert(1, "SITE_ID", sit_id_lis)
    df_raw_kpi_4G.insert(2, "CELL_ID", cell_id_lis)
  
    df_raw_kpi_4G.rename(columns={"Short name" :"Shortname" } ,inplace = True )
    df_raw_kpi_4G.fillna(value=0,inplace=True)
    PsOsPath1=os.path.join(door_path,'process_output','desired_input.xlsx')
    # df_tech.to_excel("process_output/desired_input.xlsx")
    df_raw_kpi_4G.to_excel(PsOsPath1,index=False)
    
    str_date=request.POST.get("offered_date")
    date1=datetime.datetime.strptime(str_date,"%Y-%m-%d")
    
    # date1=cal.get_date()
    dt1 = date1 - timedelta(1)
    dt2 = date1 - timedelta(2)
    dt3 = date1 - timedelta(3)
    dt4 = date1 - timedelta(4)
    dt5 = date1 - timedelta(5)
    ls=[dt1,dt2,dt3,dt4,dt5]

    def perticular_tech( tech,site_list):
        df_filtered = df_raw_kpi_4G[(df_raw_kpi_4G.SITE_ID.isin(site_list)) & (df_raw_kpi_4G.date.isin(ls)) & (df_raw_kpi_4G.Shortname.str.contains('|'.join(tech)))]
        # print(df_filtered)
        PsOs_Filtr=os.path.join(door_path,'process_output','filtered.xlsx')
        df_filtered.to_excel(PsOs_Filtr,index=False)
        df_pivoted = df_filtered.pivot_table(index=["SITE_ID","Shortname","CELL_ID"], columns="date")
        # print(df_pivoted)
        PsOs_pivot=os.path.join(door_path,'process_output','npivoted.xlsx')
        # df_pivoted.to_excel("process_output/npivoted.xlsx")
        df_pivoted.to_excel(PsOs_pivot)
        # print('---------------filter,pivot----------------',df_pivoted)
        return df_filtered,df_pivoted
    alpha = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    def num_hash(num):
            if num < 26:
                return alpha[num-1]
            else:
                q, r = num//26, num % 26
                if r == 0:
                    if q == 1:
                        return alpha[r-1]
                    else:
                        return num_hash(q-1) + alpha[r-1]
                else:
                    return num_hash(q) + alpha[r-1]
 
# Driver code

# printString(27906)

    def titleToNumber(s):

        result = 0
        for B in range(len(s)):
            result *= 26
            result += ord(s[B]) - ord('A') + 1
        return result

    def overwrite(df_pivoted,kpi_name,coln1,trend_ws):
        coln2=num_hash(titleToNumber(coln1)+1)
        coln3=num_hash(titleToNumber(coln1)+2)
        coln4=num_hash(titleToNumber(coln1)+3)
        coln5=num_hash(titleToNumber(coln1)+4)
        print(kpi_name)
        index_pivot=df_pivoted.index.to_list()
        print("index ;###############################",index_pivot)
        print(len(index_pivot))
        print("index of pivoted table: ",index_pivot)
        dr=df_pivoted[kpi_name]
        print("columns of dr table",dr.columns)
        cl=dr.columns.to_list()
        print("column list",cl)
        
        # site_id=dr["SITE_ID"].to_list() 
        # cell_id=dr["CELL_ID"].to_list()    
        col1=dr[str(cl[0])].to_list()
        col2=dr[str(cl[1])].to_list()
        col3=dr[str(cl[2])].to_list()
        col4=dr[str(cl[3])].to_list()
        col5=dr[str(cl[4])].to_list()

        trend_ws[coln1+"2"].value=cl[0]
        trend_ws[coln2+"2"].value=cl[1]
        trend_ws[coln3+"2"].value=cl[2]
        trend_ws[coln4+"2"].value=cl[3]
        trend_ws[coln5+"2"].value=cl[4]

        # me=column_index_from_string(coln5)+1
        # me=get_column_letter(me)
        for i,value in enumerate(index_pivot):
            j=i+3
            trend_ws["B"+str(j)].value=index_pivot[i][0]
            trend_ws["A"+str(j)].value=index_pivot[i][2] 
                       
            trend_ws[coln1+str(j)].value=col1[i]
            trend_ws[coln2+str(j)].value=col2[i]
            trend_ws[coln3+str(j)].value=col3[i]
            trend_ws[coln4+str(j)].value=col4[i]
            trend_ws[coln5+str(j)].value=col5[i]
                      
    def overwrite1(df_pivoted,kpi_name,coln1,trend_ws):
        print(kpi_name)
        index_pivot=df_pivoted.index.to_list()
        print("index of pivoted table: ",index_pivot)
        dr=df_pivoted[kpi_name]
        print("columns of dr table",dr.columns)
        cl=dr.columns.to_list()
        print("column list",cl)   
        col1=dr[str(cl[0])].to_list()
        trend_ws[coln1+"2"].value=cl[0]

        for i,value in enumerate(index_pivot):
            j=i+3
            trend_ws[coln1+str(j)].value=col1[i]

    def overwrite2(df_pivoted,kpi_name,coln1,trend_ws):
        print(kpi_name)
        index_pivot=df_pivoted.index.to_list()
        print("index of pivoted table: ",index_pivot)
        dr=df_pivoted[kpi_name]
        print("columns of dr table",dr.columns)
        cl=dr.columns.to_list()
        print("column list",cl)   
        col1=dr[str(cl[1])].to_list()
        trend_ws[coln1+"2"].value=cl[1]


        for i,value in enumerate(index_pivot):
            j=i+3
            trend_ws[coln1+str(j)].value=col1[i]

    def overwrite3(df_pivoted,kpi_name,coln1,trend_ws):
        print(kpi_name)
        index_pivot=df_pivoted.index.to_list()
        print("index of pivoted table: ",index_pivot)
        dr=df_pivoted[kpi_name]
        print("columns of dr table",dr.columns)
        cl=dr.columns.to_list()
        print("column list",cl)   
        col1=dr[str(cl[2])].to_list()
        trend_ws[coln1+"2"].value=cl[2]

        for i,value in enumerate(index_pivot):
            j=i+3
            trend_ws[coln1+str(j)].value=col1[i]

    # for fdd
    pivot_fdd=perticular_tech(["_F3_"],site_list)[1]
    path_of_blnk_temp=os.path.join(door_path,'templates','L1800_KPI Submission_01-Mar.xlsx')
    # path_of_blnk_temp="templates/L1800_KPI Submission_01-Mar.xlsx"
    trend_wb_L1800=openpyxl.load_workbook(path_of_blnk_temp)
    # print(trend_wb.sheetnames)
    trend_ws=trend_wb_L1800["LTE-VOLTE KPI"]
    for ta in kpii:
        if(ta=="TA Sample<500 M_Nom"):
            overwrite1(pivot_fdd,ta,"DA",trend_ws) 
            
        if(ta=="TA Sample<1 KM_Nom"):
            overwrite1(pivot_fdd,ta,"DB",trend_ws)
            
        if(ta=="TA Sample >1.5 KM_Nom_Ericsson_1"):
            overwrite1(pivot_fdd,ta,"DC",trend_ws) 

        if(ta=="TA Sample >3.4 KM_Nom"):
            overwrite1(pivot_fdd,ta,"DD",trend_ws) 

        if(ta=="TA Sample >4.45 KM_Nom"):
            overwrite1(pivot_fdd,ta,"DE",trend_ws) 
        if(ta=="TA Sample >5.7 KM_Nom"):
            overwrite1(pivot_fdd,ta,"DF",trend_ws) 

        if(ta=="TA Sample >7.2 KM_Nom"):
           overwrite1(pivot_fdd,ta,"DG",trend_ws)                     

        if(ta=="TA Sample >8.4 KM_Nom"):
            overwrite1(pivot_fdd,ta,"DH",trend_ws)
    for ta in kpii:
        if(ta=="TA Sample<500 M_Nom"):
            overwrite2(pivot_fdd,ta,"DL",trend_ws) 
            
        if(ta=="TA Sample<1 KM_Nom"):
            overwrite2(pivot_fdd,ta,"DM",trend_ws)
            
        if(ta=="TA Sample >1.5 KM_Nom_Ericsson_1"):
            overwrite2(pivot_fdd,ta,"DN",trend_ws) 

        if(ta=="TA Sample >3.4 KM_Nom"):
            overwrite2(pivot_fdd,ta,"DO",trend_ws) 

        if(ta=="TA Sample >4.45 KM_Nom"):
            overwrite2(pivot_fdd,ta,"DP",trend_ws) 
        if(ta=="TA Sample >5.7 KM_Nom"):
            overwrite2(pivot_fdd,ta,"DQ",trend_ws) 

        if(ta=="TA Sample >7.2 KM_Nom"):
           overwrite2(pivot_fdd,ta,"DR",trend_ws)                     

        if(ta=="TA Sample >8.4 KM_Nom"):
            overwrite2(pivot_fdd,ta,"DS",trend_ws)
    for ta in kpii:
        if(ta=="TA Sample<500 M_Nom"):
            overwrite3(pivot_fdd,ta,"DW",trend_ws) 
            
        if(ta=="TA Sample<1 KM_Nom"):
            overwrite3(pivot_fdd,ta,"DX",trend_ws)
            
        if(ta=="TA Sample >1.5 KM_Nom_Ericsson_1"):
            overwrite3(pivot_fdd,ta,"DY",trend_ws) 

        if(ta=="TA Sample >3.4 KM_Nom"):
            overwrite3(pivot_fdd,ta,"DZ",trend_ws) 

        if(ta=="TA Sample >4.45 KM_Nom"):
            overwrite3(pivot_fdd,ta,"EA",trend_ws) 
        if(ta=="TA Sample >5.7 KM_Nom"):
            overwrite3(pivot_fdd,ta,"EB",trend_ws) 

        if(ta=="TA Sample >7.2 KM_Nom"):
           overwrite3(pivot_fdd,ta,"EC",trend_ws)                     

        if(ta=="TA Sample >8.4 KM_Nom"):
            overwrite3(pivot_fdd,ta,"ED",trend_ws)
    for kpi_name in kpi:
            if(kpi_name=="RRC Setup Success Rate [CDBH]"):
                overwrite(pivot_fdd,kpi_name,"E",trend_ws)

            if(kpi_name=="ERAB Setup Success Rate [CDBH]"):
                overwrite(pivot_fdd,kpi_name,"K",trend_ws)

            if(kpi_name=="PS Drop Call Rate % [CDBH]"):
                overwrite(pivot_fdd,kpi_name,"Q",trend_ws)

            if(kpi_name=="MV_DL User Throughput_Mbps [CDBH]"):
                overwrite(pivot_fdd,kpi_name,"W",trend_ws)

            if(kpi_name=="MV_Average number of used DL PRBs [CDBH]"):
                overwrite(pivot_fdd,kpi_name,"AC",trend_ws)

            if(kpi_name=="MV_UL User Throughput_Mbps [CDBH]"):
                overwrite(pivot_fdd,kpi_name,"AI",trend_ws)
            
            if(kpi_name=="MV_PS handover success rate [LTE Intra System] [CDBH]"):
                overwrite(pivot_fdd,kpi_name,"AN",trend_ws)
                
            if(kpi_name=="PS handover success rate [LTE Inter System] [CDBH]"):
                overwrite(pivot_fdd,kpi_name,"AT",trend_ws)    
                
            if(kpi_name=="MV_LTE_AVG_CQI"):
                overwrite(pivot_fdd,kpi_name,"AZ",trend_ws)
            
            if(kpi_name=="VoLTE Call Setup Success rate [CBBH]"):
                overwrite(pivot_fdd,kpi_name,"BE",trend_ws)    
            
            if(kpi_name=="MV_VoLTE DCR [CBBH]"):
                overwrite(pivot_fdd,kpi_name,"BK",trend_ws) 
            
            if(kpi_name=="VoLTE IntraF HOSR Exec [CBBH]"):
                overwrite(pivot_fdd,kpi_name,"BQ",trend_ws)        
                
            if(kpi_name=="VoLTE InterF HOSR Exec [CBBH]"):
                overwrite(pivot_fdd,kpi_name,"BW",trend_ws)
                
            if(kpi_name=="MV_VoLTE Packet Loss DL [CBBH]"):
                overwrite(pivot_fdd,kpi_name,"CC",trend_ws)
                    
            if(kpi_name=="MV_VoLTE Packet Loss UL [CBBH]"):
                overwrite(pivot_fdd,kpi_name,"CI",trend_ws)
                
            if(kpi_name=="Radio NW Availability"):
                overwrite(pivot_fdd,kpi_name,"CO",trend_ws) 

            if(kpi_name=="UL RSSI [CDBH]"):
                overwrite(pivot_fdd,kpi_name,"CU",trend_ws) 
                
            
       
    # for tdd
    pivot_fdd=perticular_tech(["_T1_","_T2_"],site_list)[1]
    path_of_blnk_temp=os.path.join(door_path,'templates','TDD_ KPI Submission_01-Mar.xlsx')
    # path_of_blnk_temp="templates/TDD_ KPI Submission_01-Mar.xlsx"
    trend_wb_TDD=openpyxl.load_workbook(path_of_blnk_temp)
    # print(trend_wb.sheetnames)
    trend_ws=trend_wb_TDD["LTE-VOLTE KPI"]
    for ta in kpii:
        if(ta=="TA Sample<500 M_Nom"):
            overwrite1(pivot_fdd,ta,"DA",trend_ws) 
            
        if(ta=="TA Sample<1 KM_Nom"):
            overwrite1(pivot_fdd,ta,"DB",trend_ws)
            
        if(ta=="TA Sample >1.5 KM_Nom_Ericsson_1"):
            overwrite1(pivot_fdd,ta,"DC",trend_ws) 

        if(ta=="TA Sample >3.4 KM_Nom"):
            overwrite1(pivot_fdd,ta,"DD",trend_ws) 

        if(ta=="TA Sample >4.45 KM_Nom"):
            overwrite1(pivot_fdd,ta,"DE",trend_ws) 
        if(ta=="TA Sample >5.7 KM_Nom"):
            overwrite1(pivot_fdd,ta,"DF",trend_ws) 

        if(ta=="TA Sample >7.2 KM_Nom"):
           overwrite1(pivot_fdd,ta,"DG",trend_ws)                     

        if(ta=="TA Sample >8.4 KM_Nom"):
            overwrite1(pivot_fdd,ta,"DH",trend_ws)
    for ta in kpii:
        if(ta=="TA Sample<500 M_Nom"):
            overwrite2(pivot_fdd,ta,"DL",trend_ws) 
            
        if(ta=="TA Sample<1 KM_Nom"):
            overwrite2(pivot_fdd,ta,"DM",trend_ws)
            
        if(ta=="TA Sample >1.5 KM_Nom_Ericsson_1"):
            overwrite2(pivot_fdd,ta,"DN",trend_ws) 

        if(ta=="TA Sample >3.4 KM_Nom"):
            overwrite2(pivot_fdd,ta,"DO",trend_ws) 

        if(ta=="TA Sample >4.45 KM_Nom"):
            overwrite2(pivot_fdd,ta,"DP",trend_ws) 
        if(ta=="TA Sample >5.7 KM_Nom"):
            overwrite2(pivot_fdd,ta,"DQ",trend_ws) 

        if(ta=="TA Sample >7.2 KM_Nom"):
           overwrite2(pivot_fdd,ta,"DR",trend_ws)                     

        if(ta=="TA Sample >8.4 KM_Nom"):
            overwrite2(pivot_fdd,ta,"DS",trend_ws)
    for ta in kpii:
        if(ta=="TA Sample<500 M_Nom"):
            overwrite3(pivot_fdd,ta,"DW",trend_ws) 
            
        if(ta=="TA Sample<1 KM_Nom"):
            overwrite3(pivot_fdd,ta,"DX",trend_ws)
            
        if(ta=="TA Sample >1.5 KM_Nom_Ericsson_1"):
            overwrite3(pivot_fdd,ta,"DY",trend_ws) 

        if(ta=="TA Sample >3.4 KM_Nom"):
            overwrite3(pivot_fdd,ta,"DZ",trend_ws) 

        if(ta=="TA Sample >4.45 KM_Nom"):
            overwrite3(pivot_fdd,ta,"EA",trend_ws) 
        if(ta=="TA Sample >5.7 KM_Nom"):
            overwrite3(pivot_fdd,ta,"EB",trend_ws) 

        if(ta=="TA Sample >7.2 KM_Nom"):
           overwrite3(pivot_fdd,ta,"EC",trend_ws)                     

        if(ta=="TA Sample >8.4 KM_Nom"):
            overwrite3(pivot_fdd,ta,"ED",trend_ws)




    for kpi_name in kpi:
            if(kpi_name=="RRC Setup Success Rate [CDBH]"):
                overwrite(pivot_fdd,kpi_name,"E",trend_ws)

            if(kpi_name=="ERAB Setup Success Rate [CDBH]"):
                overwrite(pivot_fdd,kpi_name,"K",trend_ws)

            if(kpi_name=="PS Drop Call Rate % [CDBH]"):
                overwrite(pivot_fdd,kpi_name,"Q",trend_ws)

            if(kpi_name=="MV_DL User Throughput_Mbps [CDBH]"):
                overwrite(pivot_fdd,kpi_name,"W",trend_ws)

            if(kpi_name=="MV_Average number of used DL PRBs [CDBH]"):
                overwrite(pivot_fdd,kpi_name,"AC",trend_ws)

            if(kpi_name=="MV_UL User Throughput_Mbps [CDBH]"):
                overwrite(pivot_fdd,kpi_name,"AI",trend_ws)
            
            if(kpi_name=="MV_PS handover success rate [LTE Intra System] [CDBH]"):
                overwrite(pivot_fdd,kpi_name,"AN",trend_ws)
                
            if(kpi_name=="PS handover success rate [LTE Inter System] [CDBH]"):
                overwrite(pivot_fdd,kpi_name,"AT",trend_ws)    
                
            if(kpi_name=="MV_LTE_AVG_CQI"):
                overwrite(pivot_fdd,kpi_name,"AZ",trend_ws)
            
            if(kpi_name=="VoLTE Call Setup Success rate [CBBH]"):
                overwrite(pivot_fdd,kpi_name,"BE",trend_ws)    
            
            if(kpi_name=="MV_VoLTE DCR [CBBH]"):
                overwrite(pivot_fdd,kpi_name,"BK",trend_ws) 
            
            if(kpi_name=="VoLTE IntraF HOSR Exec [CBBH]"):
                overwrite(pivot_fdd,kpi_name,"BQ",trend_ws)        
                
            if(kpi_name=="VoLTE InterF HOSR Exec [CBBH]"):
                overwrite(pivot_fdd,kpi_name,"BW",trend_ws)
                
            if(kpi_name=="MV_VoLTE Packet Loss DL [CBBH]"):
                overwrite(pivot_fdd,kpi_name,"CC",trend_ws)
                    
            if(kpi_name=="MV_VoLTE Packet Loss UL [CBBH]"):
                overwrite(pivot_fdd,kpi_name,"CI",trend_ws)
                
            if(kpi_name=="Radio NW Availability"):
                overwrite(pivot_fdd,kpi_name,"CO",trend_ws) 

            if(kpi_name=="UL RSSI [CDBH]"):
                overwrite(pivot_fdd,kpi_name,"CU",trend_ws) 
                
            
    pivot_fdd=perticular_tech(["_F8_"],site_list)[1]
    path_of_blnk_temp=os.path.join(door_path,'templates','L900_KPI Submission _01-Mar.xlsx')
    # path_of_blnk_temp="templates/L900_KPI Submission _01-Mar.xlsx"
    trend_wb_L900=openpyxl.load_workbook(path_of_blnk_temp)
    trend_ws=trend_wb_L900["L900 KPI"]
    for ta in kpii:
        if(ta=="TA Sample<500 M_Nom"):
            overwrite1(pivot_fdd,ta,"DA",trend_ws) 
            
        if(ta=="TA Sample<1 KM_Nom"):
            overwrite1(pivot_fdd,ta,"DB",trend_ws)
            
        if(ta=="TA Sample >1.5 KM_Nom_Ericsson_1"):
            overwrite1(pivot_fdd,ta,"DC",trend_ws) 

        if(ta=="TA Sample >3.4 KM_Nom"):
            overwrite1(pivot_fdd,ta,"DD",trend_ws) 

        if(ta=="TA Sample >4.45 KM_Nom"):
            overwrite1(pivot_fdd,ta,"DE",trend_ws) 
        if(ta=="TA Sample >5.7 KM_Nom"):
            overwrite1(pivot_fdd,ta,"DF",trend_ws) 

        if(ta=="TA Sample >7.2 KM_Nom"):
           overwrite1(pivot_fdd,ta,"DG",trend_ws)                     

        if(ta=="TA Sample >8.4 KM_Nom"):
            overwrite1(pivot_fdd,ta,"DH",trend_ws)
    for ta in kpii:
        if(ta=="TA Sample<500 M_Nom"):
            overwrite2(pivot_fdd,ta,"DL",trend_ws) 
            
        if(ta=="TA Sample<1 KM_Nom"):
            overwrite2(pivot_fdd,ta,"DM",trend_ws)
            
        if(ta=="TA Sample >1.5 KM_Nom_Ericsson_1"):
            overwrite2(pivot_fdd,ta,"DN",trend_ws) 

        if(ta=="TA Sample >3.4 KM_Nom"):
            overwrite2(pivot_fdd,ta,"DO",trend_ws) 

        if(ta=="TA Sample >4.45 KM_Nom"):
            overwrite2(pivot_fdd,ta,"DP",trend_ws) 
        if(ta=="TA Sample >5.7 KM_Nom"):
            overwrite2(pivot_fdd,ta,"DQ",trend_ws) 

        if(ta=="TA Sample >7.2 KM_Nom"):
           overwrite2(pivot_fdd,ta,"DR",trend_ws)                     

        if(ta=="TA Sample >8.4 KM_Nom"):
            overwrite2(pivot_fdd,ta,"DS",trend_ws)
    for ta in kpii:
        if(ta=="TA Sample<500 M_Nom"):
            overwrite3(pivot_fdd,ta,"DW",trend_ws) 
            
        if(ta=="TA Sample<1 KM_Nom"):
            overwrite3(pivot_fdd,ta,"DX",trend_ws)
            
        if(ta=="TA Sample >1.5 KM_Nom_Ericsson_1"):
            overwrite3(pivot_fdd,ta,"DY",trend_ws) 

        if(ta=="TA Sample >3.4 KM_Nom"):
            overwrite3(pivot_fdd,ta,"DZ",trend_ws) 

        if(ta=="TA Sample >4.45 KM_Nom"):
            overwrite3(pivot_fdd,ta,"EA",trend_ws) 
        if(ta=="TA Sample >5.7 KM_Nom"):
            overwrite3(pivot_fdd,ta,"EB",trend_ws) 

        if(ta=="TA Sample >7.2 KM_Nom"):
           overwrite3(pivot_fdd,ta,"EC",trend_ws)                     

        if(ta=="TA Sample >8.4 KM_Nom"):
            overwrite3(pivot_fdd,ta,"ED",trend_ws)

    for kpi_name in kpi:
            if(kpi_name=="RRC Setup Success Rate [CDBH]"):
                overwrite(pivot_fdd,kpi_name,"E",trend_ws)

            if(kpi_name=="ERAB Setup Success Rate [CDBH]"):
                overwrite(pivot_fdd,kpi_name,"K",trend_ws)

            if(kpi_name=="PS Drop Call Rate % [CDBH]"):
                overwrite(pivot_fdd,kpi_name,"Q",trend_ws)

            if(kpi_name=="MV_DL User Throughput_Mbps [CDBH]"):
                overwrite(pivot_fdd,kpi_name,"W",trend_ws)

            if(kpi_name=="MV_Average number of used DL PRBs [CDBH]"):
                overwrite(pivot_fdd,kpi_name,"AC",trend_ws)

            if(kpi_name=="MV_UL User Throughput_Mbps [CDBH]"):
                overwrite(pivot_fdd,kpi_name,"AI",trend_ws)
            
            if(kpi_name=="MV_PS handover success rate [LTE Intra System] [CDBH]"):
                overwrite(pivot_fdd,kpi_name,"AN",trend_ws)
                
            if(kpi_name=="PS handover success rate [LTE Inter System] [CDBH]"):
                overwrite(pivot_fdd,kpi_name,"AT",trend_ws)    
                
            if(kpi_name=="MV_LTE_AVG_CQI"):
                overwrite(pivot_fdd,kpi_name,"AZ",trend_ws)
            
            if(kpi_name=="VoLTE Call Setup Success rate [CBBH]"):
                overwrite(pivot_fdd,kpi_name,"BE",trend_ws)    
            
            if(kpi_name=="MV_VoLTE DCR [CBBH]"):
                overwrite(pivot_fdd,kpi_name,"BK",trend_ws) 
            
            if(kpi_name=="VoLTE IntraF HOSR Exec [CBBH]"):
                overwrite(pivot_fdd,kpi_name,"BQ",trend_ws)        
                
            if(kpi_name=="VoLTE InterF HOSR Exec [CBBH]"):
                overwrite(pivot_fdd,kpi_name,"BW",trend_ws)
                
            if(kpi_name=="MV_VoLTE Packet Loss DL [CBBH]"):
                overwrite(pivot_fdd,kpi_name,"CC",trend_ws)
                    
            if(kpi_name=="MV_VoLTE Packet Loss UL [CBBH]"):
                overwrite(pivot_fdd,kpi_name,"CI",trend_ws)
                
            if(kpi_name=="Radio NW Availability"):
                overwrite(pivot_fdd,kpi_name,"CO",trend_ws) 

            if(kpi_name=="UL RSSI [CDBH]"):
                overwrite(pivot_fdd,kpi_name,"CU",trend_ws) 
    save_output_L900=os.path.join(door_path,'output','toBeZipped','22april_L900_KPI_Submission_01Mar.xlsx')            
    # trend_wb_L900.save("output/22april_L900_KPI Submission _01Mar.xlsx")
    trend_wb_L900.save(save_output_L900)
    save_output_L18=os.path.join(door_path,'output','toBeZipped','22april_L1800_KPI_Submission_01Mar.xlsx')            
    trend_wb_L1800.save(save_output_L18)
    save_output_tdd=os.path.join(door_path,'output','toBeZipped','22april_TDD_KPI_Submission_01Mar.xlsx') 
    trend_wb_TDD.save(save_output_tdd)
    print('successfully')
    
      
    def get_all_file_paths(directory):
    
        # initializing empty file paths list
        file_paths = []
    
        # crawling through directory and subdirectories
        for root, directories, files in os.walk(directory):
            for filename in files:
                # join the two strings in order to form the full filepath.
                filepath = os.path.join(root, filename)
                file_paths.append(filepath)
    
        # returning all file paths
        return file_paths        
    
    def main1():
        # path to folder which needs to be zipped
        directory = os.path.join(door_path,'output','toBeZipped')
    
        # calling function to get all file paths in the directory
        file_paths = get_all_file_paths(directory)
    
        # printing the list of all files to be zipped
        print('Following files will be zipped:')
        for file_name in file_paths:
            print(file_name)
    
        # writing files to a zipfile
        with ZipFile(os.path.join(door_path,'output','output','Ap_output.zip'),'w') as zip:
            # writing each file one by one
            for file in file_paths:
                zip.write(file)
    
        print('All files zipped successfully!')        
    
    
    
    main1()
    # download_path1=os.path.join(MEDIA_URL,'trends','ap',"ap4G",'output','22april_L900_KPI_Submission_01Mar.xlsx')
    # download_path2=os.path.join(MEDIA_URL,'trends','ap',"ap4G",'output','22april_L1800_KPI_Submission_01Mar.xlsx')
    # download_path3=os.path.join(MEDIA_URL,'trends','ap',"ap4G","hryanatrend",'output','22april_TDD_KPI_Submission_01Mar.xlsx')
    download_path=os.path.join(MEDIA_URL,'trends','ap',"ap4G",'output','output','Ap_output.zip')
    # return Response({'status':True,'Download_url':download_path1,'Download_url2':download_path2,'Download_url3':download_path3})
    return Response({'status':True,'message':'successfully','Download_url':download_path})

#################################################### GSM ###########################################
@api_view(["POST"])
def old_ap2G_trend(request):
    raw_kpi_2G=request.FILES['raw_kpi_2G'] if 'raw_kpi_2G' in request.FILES else None
    if raw_kpi_2G:
        location=MEDIA_ROOT +r'\trends\temporary_files'
        fs=FileSystemStorage(location=location)
        file_2G=fs.save(raw_kpi_2G.name,raw_kpi_2G)
        file_2G_path=fs.path(file_2G)
        df_raw_kpi_2G=pd.read_excel(file_2G_path)
        print(df_raw_kpi_2G)
        os.remove(path=file_2G_path)

    site_list_2G=request.FILES['site_list_2G'] if 'site_list_2G' in request.FILES else None
    if site_list_2G:
        location=MEDIA_ROOT+r'\trends\temporary_files'
        fs=FileSystemStorage(location=location)
        file_2G_sheet=fs.save(site_list_2G.name,site_list_2G)
        file_2G_path=fs.path(file_2G_sheet)
        df_2G_site=pd.read_excel(file_2G_path)
        print(df_2G_site)
        os.remove(path=file_2G_path) 


   
    door_path=os.path.join(MEDIA_ROOT,'trends','ap','ap2G')
    # save_path=os.path.join(door_path,'input','GSM_AP.xlsx')
    # df_raw_kpi_2G = pd.read_excel(save_path)
    STR=os.path.join(door_path,'templates','G1800_KPI_01-Mar.xlsx')
    # STR = 'templates/G1800_KPI_01-Mar.xlsx'
    wb = openpyxl.load_workbook(STR)
    ws = wb['2G KPI']
    print(df_raw_kpi_2G)


      

    df_raw_kpi_2G["Short name"] = df_raw_kpi_2G["Short name"].fillna(method="ffill")  #######for (forward fill)ffill is used to copy  and fill
  


    df_raw_kpi_2G.columns.values[1] = 'Date'  #####for empty colum
    
    

    A = []
    B = []

    for x in df_raw_kpi_2G['Short name']:
        if(" " in str(x)):
            site_id = x.split(" ")[0]
            A.append(site_id)
        else:
            site_id = str(x)[:-1]
            A.append(site_id)
        if(" " in str(x)):
            cell_id = x.split(" ")[-1][:6]
            B.append(cell_id )
        else:
            cell_id  = str(x)[:7]
            B.append(cell_id )
    df_raw_kpi_2G.insert(0,'SITE_ID',A)
    df_raw_kpi_2G.insert(4,'cell_id',B)

    PsOsPath=os.path.join(door_path,'process_output','final.xlsx')
    df_raw_kpi_2G.to_excel(PsOsPath,index=False)



    excel_file_1 = PsOsPath
    # PsOsPath2=os.path.join(door_path,'input','SITES2G.xlsx')
    # excel_file_2 = PsOsPath2

    df1 = pd.read_excel(excel_file_1)
    # df2 = pd.read_excel(excel_file_2)

#     # df2 = pd.read_excel(excel_file_2)

    df1.rename(columns={"Site ID": "SITE_ID"}, inplace=True)
    df_2G_site.rename(columns={"Site ID": "SITE_ID"}, inplace=True)

    g2_kpi=["SDCCH Drop Call Rate","SDCCH Blocking Rate","TCH Blocking Rate","Drop Call Rate","RX Quality [BBH]","Handover Success Rate",	
        "Handover Success Rate_Denom","Handover Success Rate_Nom","TCH Assignment Success Rate","Network availability [RNA]","Total Voice Traffic",]


    filtered_df_1 = df1[(df1.SITE_ID.isin(list(df_2G_site['SITE_ID'])))]

    print(filtered_df_1)
    PsOs_filtr=os.path.join(door_path,'process_output','filtered_df_1.xlsx')
    filtered_df_1.to_excel(PsOs_filtr,index=False)
    df1 = pd.read_excel(PsOs_filtr)

    df_pivot = df1.pivot_table(columns='Date', index=[ 'SITE_ID',"cell_id" ])
    print('df_pivot')
   
    PsOs_pivot=os.path.join(door_path,'process_output','pivot.xlsx')
    df_pivot.to_excel(PsOs_pivot)


    alpha = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    def num_hash(num):
            if num < 26:
                return alpha[num-1]
            else:
                q, r = num//26, num % 26
                if r == 0:
                    if q == 1:
                        return alpha[r-1]
                    else:
                        return num_hash(q-1) + alpha[r-1]
                else:
                    return num_hash(q) + alpha[r-1]



    def titleToNumber(s):
        # This process is similar to binary-to-
        # decimal conversion
            result = 0
            for B in range(len(s)):
                result *= 26
                result += ord(s[B]) - ord('A') + 1
            return result

    str_date=request.POST.get("offered_date")
    date1=datetime.datetime.strptime(str_date,"%Y-%m-%d")
    d1=date1-timedelta(1)
    d2=date1-timedelta(2)
    d3=date1-timedelta(3)
    d4=date1-timedelta(4)
    d5=date1-timedelta(5)
    cl=[d1,d2,d3,d4,d5]
    # index=df_pivot.index
    

    def overwrite(kpi_name,coln1):
        coln2=num_hash(titleToNumber(coln1)+1)
        coln3=num_hash(titleToNumber(coln1)+2)
        coln4=num_hash(titleToNumber(coln1)+3)
        coln5=num_hash(titleToNumber(coln1)+4)
        print(kpi_name)
        index=df_pivot.index
        # print("index ;###############################",index_pivot)
        # print(len(index_pivot))
        print("doooooonnenn")
        dr=df_pivot[kpi_name]
        print("columns of dr tab")
        li=dr.columns
        print("column list")


        col1=dr[li[0]].to_list()
        col2=dr[li[1]].to_list()
        col3=dr[li[2]].to_list()
        col4=dr[li[3]].to_list()
        col5=dr[li[4]].to_list()


        ws[coln1+"2"].value=cl[4]
        ws[coln2+"2"].value=cl[3]
        ws[coln3+"2"].value=cl[2]
        ws[coln4+"2"].value=cl[1]
        ws[coln5+"2"].value=cl[0]

        for i,value in enumerate(index):
            j=i+6
              
            ws['A'+str(j)].value=index[i][0]
            ws['B'+str(j)].value=index[i][1]



            ws[coln1+str(j)].value=col1[i]
            ws[coln2+str(j)].value=col2[i]
            ws[coln3+str(j)].value=col3[i]
            ws[coln4+str(j)].value=col4[i]
            ws[coln5+str(j)].value=col5[i]


    for kpi_name in g2_kpi:
            if(kpi_name=='SDCCH Drop Call Rate'):
                overwrite(kpi_name,'E')

            if(kpi_name=='SDCCH Blocking Rate'):
                overwrite(kpi_name,'K')
           
            if(kpi_name=='TCH Blocking Rate'):
                overwrite(kpi_name,'Q')

            if(kpi_name=='Drop Call Rate'):
                overwrite(kpi_name,'W')

            if(kpi_name=='RX Quality [BBH]'):
                overwrite(kpi_name,'AC') 
                
            if(kpi_name=='Handover Success Rate'):
                overwrite(kpi_name,'AI') 

            if(kpi_name=='Handover Success Rate_Denom'):
                overwrite(kpi_name,'AO') 

            if(kpi_name=='Handover Success Rate_Nom'):
                overwrite(kpi_name,'AT') 

            if(kpi_name=='TCH Assignment Success Rate'):
                overwrite(kpi_name,'BD') 

            if(kpi_name=='Network availability [RNA]'):
                overwrite(kpi_name,'BJ')    

            if(kpi_name=='Total Voice Traffic'):
                overwrite(kpi_name,'BO') 

    SaveOutput=os.path.join(MEDIA_ROOT,'trends','ap','ap2G','output','gsm1.xlsx')
    wb.save(SaveOutput)
  
    download_path=os.path.join(MEDIA_URL,'trends','ap','ap2G','output','gsm1.xlsx')
    return Response({"status":True,'Download_url':download_path,"message":"successfully uploaded"})

##############--------------------------------------------------------------------##################################     


