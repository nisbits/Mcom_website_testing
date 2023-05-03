from django.shortcuts import render
from django.http import HttpResponse
from mcom_website.settings import MEDIA_ROOT,MEDIA_URL
from rest_framework.decorators import api_view
from rest_framework.response import Response
from tkinter import *
from django.core.files.storage import FileSystemStorage
from tkinter import filedialog, messagebox, ttk
import openpyxl
import pandas as pd
import numpy as np
from datetime import date, timedelta
import datetime

import os

@api_view(["POST"])
def old_bih_trend(request):
    raw_kpi_4G=request.FILES['raw_kpi'] if "raw_kpi" in request.FILES else None
    if raw_kpi_4G:
        location=MEDIA_ROOT+r'\trends\temporary_files'
        fs=FileSystemStorage(location=location)
        file=fs.save(raw_kpi_4G.name,raw_kpi_4G)
        file_path=fs.path(file)
        df_raw_kpi=pd.read_excel(file_path)
        print('------------------------uuuuu---------------')

        print(df_raw_kpi)
        os.remove(path=file_path)

    site_list_4G=request.FILES["site_list"] if "site_list" in request.FILES else None
    if site_list_4G:
        location=MEDIA_ROOT+r'\trends\temporary_files'
        fs=FileSystemStorage(location=location)
        file=fs.save(site_list_4G.name,site_list_4G)
        file_site_path=fs.path(file)
        df_site_list=pd.read_excel(file_site_path)
        print(df_site_list)
        os.remove(path=file_site_path)

    door_path=os.path.join(MEDIA_ROOT,'trends','bihar')
   
    df_raw_kpi['Short name']=df_raw_kpi['Short name'].fillna(method='ffill')
    print('_________date_________',df_raw_kpi)
    df_raw_kpi.columns.values[1]='Date'
    
    df_raw_kpi['DL User Throughput_Kbps [CDBH]']=(df_raw_kpi['DL User Throughput_Kbps [CDBH]']/1024)
    df_raw_kpi.columns.values[21]='DL User Throughput_Mbps [CDBH]'

    split=[]
    for cell in df_raw_kpi['Short name']:
        if('_' in cell):
            id=cell.split('_')[-2][:-1]
            split.append(id)
        else:
            id=cell[:-1]
            split.append(id)
    df_raw_kpi.insert(0,'site id',split)

    techlist=[]
    for sn in df_raw_kpi['Short name']:
            if('_' in sn):
                site_id = sn.split("_")[-2][:-1]
            
            else:
                site_id=sn[:-1]
            if('_F1_' in sn or '_F2_' in sn or '_F3_' in sn or '_F8_' in sn  or '_T1' in sn or '_T2_' in sn ):  
                if('_F1_' in sn or '_F2_' in sn or '_F3_' in sn or '_F8_' in sn ):
                    tech='FD-' + site_id
                    techlist.append(tech)
                if('_T1_' in sn or '_T2_' in sn):
                    tech='TD-' + site_id
                    techlist.append(tech)
                        
            else:
                tech=site_id
                techlist.append(tech)
    df_raw_kpi.insert(2,'site type',techlist)

    ecgi=[]
    for cell1 in df_raw_kpi['4G_ECGI']:
        if('-' in cell1):
            lnbts_id = cell1.split("-")[-2]
            ecgi.append(lnbts_id)
        
        else:
            lnbts_id=cell1[:-1]
            ecgi.append(lnbts_id)  
        
    df_raw_kpi.insert(4,'LNBTS ID', ecgi)  
    print(df_raw_kpi)

    ecgi=[]
    for cell2 in df_raw_kpi['4G_ECGI']:
        if('-' in cell2):
            lnbts_Name = cell2.split("-")[-1]
            ecgi.append(lnbts_Name)
        
        else:
            lnbts_Name=cell2[:-1]
            ecgi.append(lnbts_Name)  
        
    df_raw_kpi.insert(5,'ss', ecgi)  
    print(df_raw_kpi)

    missing_ecgi=[]   
    for cell3 in df_raw_kpi['4G_ECGI']:
        if('---' in cell3):
            if('-' in cell3):
                sector1 = cell3.split("-")[-1]
                missing_ecgi.append(sector1)
        
        else:
            sector1=cell3.split('-')[-1][-1]
            missing_ecgi.append(sector1)      
    df_raw_kpi.insert(6,'sector',missing_ecgi)  
    print(df_raw_kpi)

    df_raw_kpi['SECTOR']=df_raw_kpi['LNBTS ID'].astype(str) + '-' + df_raw_kpi['ss']
    df_raw_kpi

    df_raw_kpi.fillna(value=0,inplace=True)  
    PsOsPath1=os.path.join(door_path,'process output','fill.xlsx')
    
    df_raw_kpi.to_excel(PsOsPath1,index=False)
    
    excelfile_1=PsOsPath1
    
    # PsOsPath2=os.path.join(door_path,'project file','site.xlsx')
    
    # excelfile_2=PsOsPath2
    df1=pd.read_excel(excelfile_1)
    # df2=pd.read_excel(excelfile_2)

    df1.rename(columns={'site id':'site_id'},inplace=True)
    # df_site_list.rename(columns={'site id':'site_id'},inplace=True)

    kpi=['Radio NW Availability','E-UTRAN Average CQI [CDBH]','UL User Throughput_Kbps [CDBH]','VoLTE Traffic_24Hrs',
        'UL PUCCH SINR [CDBH]','Data Volume DL - Total [MB] [CDBH]','RRC Setup Success Rate [CDBH]','ERAB Setup Success Rate [CDBH]','VoLTE DCR_Nom [CBBH]',
        'VoLTE DCR_Denom [CBBH]','VoLTE DCR [CBBH]','PS handover success rate [LTE Intra System] [CDBH]','VoLTE SRVCC SR_Nom [CBBH]','VoLTE SRVCC SR_Denom [CBBH]',
        'VoLTE SRVCC SR','VoLTE ERAB Setup Success Rate [CBBH]','VoLTE SRVCC SR_Denom',
        'DL User Throughput_Mbps [CDBH]','Average number of used DL PRBs [CDBH]','Average UE Distance_KM [CDBH]','VoLTE Packet Loss DL [CBBH]',
        'VoLTE Packet Loss UL [CBBH]','4G Data Volume_L1800 [GB]','4G Data Volume_L900 [GB]','MV_4G Data Volume_GB_TDD','4G Data Volume_L2100 [GB]','4G Data Volume [GB]',
        'PS Drop Call Rate % [CDBH]']

    filtered_df_1 = df1[(df1.site_id.isin(list(df_site_list['2G ID'])))]
    
    print(filtered_df_1)
    PsOsPath=os.path.join(door_path,'process output','filtered1.xlsx')
    filtered_df_1.to_excel(PsOsPath,index=False)       
    # print(df)
    
    df1 = pd.read_excel(PsOsPath)
    df_pivot = df1.pivot_table(values=kpi, columns='Date', index=['Short name', 'site_id','site type','LNBTS ID','ss','SECTOR','4G_ECGI','sector'])

    # df.fillna(value=0,inplace=True)
    PsOsPathpivot=os.path.join(door_path,'process output','pivot.xlsx')
    df_pivot.to_excel(PsOsPathpivot)
    

    STR=os.path.join(door_path,'templates','trend bihar.xlsx')
    wb=openpyxl.load_workbook(STR)
    ws=wb.active
    alpha = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    def num_hash(num):
            if num < 26:
                return alpha[num-1]
            else:
                q, r = num//26, num % 26
                if r == 0:
                    if q == 1:
                        return alpha[r-1]
                    else:
                        return num_hash(q-1) + alpha[r-1]
                else:
                    return num_hash(q) + alpha[r-1]
            
            
    def titleToNumber(s):
        result= 0
        for B in range(len(s)):
            result *= 26
            result += ord(s[B]) - ord('A') + 1
        return result
    str_date=request.POST.get("offered_date")
    date1 =datetime.datetime.strptime(str_date,'%Y-%m-%d')
    d1=date1-timedelta(1)
    d2=date1-timedelta(2)
    d3=date1-timedelta(3)
    d4=date1-timedelta(4)
    d5=date1-timedelta(5)
    cl=[d1,d2,d3,d4,d5]
    def overwrite(kpi_name,coln1):
        coln2=num_hash(titleToNumber(coln1)+1)
        coln3=num_hash(titleToNumber(coln1)+2)
        coln4=num_hash(titleToNumber(coln1)+3)
        coln5=num_hash(titleToNumber(coln1)+4)
        print(kpi_name)
        index=df_pivot.index
        print('donnnnnnnnnnnnnnnnne')
        print(len(index))
        dr=df_pivot[kpi_name]
        li=dr.columns
        col1=dr[li[0]].to_list()
        col2=dr[li[1]].to_list()
        col3=dr[li[2]].to_list()
        col4=dr[li[3]].to_list()
        col5=dr[li[4]].to_list()

        ws[coln1+"2"].value=cl[4]
        ws[coln2+"2"].value=cl[3]
        ws[coln3+"2"].value=cl[2]
        ws[coln4+"2"].value=cl[1]
        ws[coln5+"2"].value=cl[0]

        for i,value in enumerate(index):
                j=i+3
                ws['C'+str(j)].value='Relocation NT'
                
                ws['H'+str(j)].value=date1
           
                ws['A'+str(j)].value=index[i][1]
                ws['B'+str(j)].value=index[i][2]
                ws['D'+str(j)].value=index[i][3]
                ws['F'+str(j)].value=index[i][7]
                ws['I'+str(j)].value=index[i][7]
                ws['E'+str(j)].value=index[i][5]
                ws['K'+str(j)].value=index[i][4]
                ws['L'+str(j)].value=index[i][6]
                
                ws['G'+str(j)].value=index[i][1]
                ws['J'+str(j)].value=index[i][0]
               
                ws[coln1+str(j)].value=col1[i]
                ws[coln2+str(j)].value=col2[i]
                ws[coln3+str(j)].value=col3[i]
                ws[coln4+str(j)].value=col4[i]
                ws[coln5+str(j)].value=col5[i]
                
    for kpi_name in kpi:
        if(kpi_name=='4G Data Volume [GB]'):
            overwrite(kpi_name,'R') 
            
        if(kpi_name=='Radio NW Availability'):
            overwrite(kpi_name,'X')    
            
        if(kpi_name=='E-UTRAN Average CQI [CDBH]'):
            overwrite(kpi_name,'AD')   
            
        if(kpi_name=='UL User Throughput_Kbps [CDBH]'):
            overwrite(kpi_name,'AJ')
            
        if(kpi_name=='VoLTE Traffic_24Hrs'):
            overwrite(kpi_name,'AP')
            
        if(kpi_name=='UL PUCCH SINR [CDBH]'):
            overwrite(kpi_name,'AV')
            
        if(kpi_name=='Data Volume DL - Total [MB] [CDBH]'):
            overwrite(kpi_name,'BB')  
            
        if(kpi_name=='RRC Setup Success Rate [CDBH]'):
            overwrite(kpi_name,'BH')
            
        if(kpi_name=='ERAB Setup Success Rate [CDBH]'):
            overwrite(kpi_name,'BN') 
            
        if(kpi_name=='VoLTE DCR_Nom [CBBH]'):
            overwrite(kpi_name,'BT') 
            
        if(kpi_name=='VoLTE DCR_Denom [CBBH]'):
            overwrite(kpi_name,'BZ')
            
        if(kpi_name=='VoLTE DCR [CBBH]'):
            overwrite(kpi_name,'CF')
            
        if(kpi_name=='PS handover success rate [LTE Intra System] [CDBH]'):
            overwrite(kpi_name,'CL')
            
        if(kpi_name=='VoLTE SRVCC SR_Nom [CBBH]'):
            overwrite(kpi_name,'CR')
            
        if(kpi_name=='VoLTE SRVCC SR_Denom [CBBH]'):
            overwrite(kpi_name,'CX')
            
        # if(kpi_name=='VoLTE SRVCC SR'):
        #     overwrite(kpi_name,'DD')
            
        if(kpi_name=='VoLTE ERAB Setup Success Rate [CBBH]'):
            overwrite(kpi_name,'DJ')
            
        if(kpi_name=='VoLTE SRVCC SR_Denom'):
            overwrite(kpi_name,'DP')
            
        if(kpi_name=='VoLTE SRVCC SR [CBBH]'):
            overwrite(kpi_name,'DV') 
            
        if(kpi_name=='DL User Throughput_Mbps [CDBH]'):
            overwrite(kpi_name,'EB')
            
        if(kpi_name=='Average number of used DL PRBs [CDBH]'):
            overwrite(kpi_name,'EH')
            
        if(kpi_name=='Average UE Distance_KM [CDBH]'):
            overwrite(kpi_name,'EN')
            
        if(kpi_name=='VoLTE Packet Loss DL [CBBH]'):
            overwrite(kpi_name,'ET')
            
        if(kpi_name=='VoLTE Packet Loss UL [CBBH]'):
            overwrite(kpi_name,'EZ')
            
        if(kpi_name=='4G Data Volume_L1800 [GB]'):
            overwrite(kpi_name,'FF') 
            
        if(kpi_name=='4G Data Volume_L900 [GB]'):
            overwrite(kpi_name,'FL')
            
            
        if(kpi_name=='MV_4G Data Volume_GB_TDD'):
            overwrite(kpi_name,'FR')
            
        if(kpi_name=='4G Data Volume_L2100 [GB]'):
            overwrite(kpi_name,'GD')
            
        if(kpi_name=='4G Data Volume [GB]'):
            overwrite(kpi_name,'GJ')
            
        if(kpi_name=='PS Drop Call Rate % [CDBH]'):
            overwrite(kpi_name,'GP')            
    SaveOutput=os.path.join(door_path,'output','biharoutput.xlsx')        
    wb.save(SaveOutput) 
    download_path=os.path.join(MEDIA_URL,'trends','bihtrend','output','biharoutput.xlsx')
    return Response({"status":True,'message':'successfully','Download_url':download_path}) 
 
############################################################################################