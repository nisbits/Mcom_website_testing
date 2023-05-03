from django.shortcuts import render
import pandas as pd
import datetime
from datetime import date, timedelta
from openpyxl import Workbook,load_workbook
from openpyxl.utils import  get_column_letter
import openpyxl
from tkinter import *
import tkinter as tk
from tkinter import filedialog,messagebox,ttk
from mcom_website.settings import MEDIA_ROOT,MEDIA_URL
from django.core.files.storage import FileSystemStorage
import os
from datetime import date, timedelta
import glob

import glob
import openpyxl
from rest_framework.decorators import api_view
from rest_framework.response import Response

@api_view(["POST"])
def old_pb_trend(request):
    raw_kpi_4G=request.FILES['raw_kpi_4G'] if 'raw_kpi_4G' in request.FILES else None
    if raw_kpi_4G:
        location=MEDIA_ROOT +r'\trends\temporary_files'
        fs=FileSystemStorage(location=location)
        file=fs.save(raw_kpi_4G.name,raw_kpi_4G)
        file_path=fs.path(file)
        df_raw_kpi_4G=pd.read_excel(file_path)
        print(df_raw_kpi_4G)
        os.remove(path=file_path)

    site_list_4G=request.FILES['site_list_4G'] if 'site_list_4G' in request.FILES else None
    if site_list_4G:
        location=MEDIA_ROOT+r'\trends\temporary_files'
        fs=FileSystemStorage(location=location)
        file_sheet=fs.save(site_list_4G.name,site_list_4G)
        file_path=fs.path(file_sheet)
        df_site_4G=pd.read_excel(file_path)
        print(df_site_4G)
        os.remove(path=file_path)

    raw_kpi_2G=request.FILES['2G_raw_kpi'] if '2G_raw_kpi' in request.FILES else None
    if raw_kpi_2G:
        location=MEDIA_ROOT +r'\trends\temporary_files'
        fs=FileSystemStorage(location=location)
        file_2G=fs.save(raw_kpi_2G.name,raw_kpi_2G)
        file_2G_path=fs.path(file_2G)
        df_raw_kpi_2G=pd.read_excel(file_2G_path)
        print(df_raw_kpi_2G)
        os.remove(path=file_2G_path)

    site_list_4G=request.FILES['site_list_2G'] if 'site_list_2G' in request.FILES else None
    if site_list_4G:
        location=MEDIA_ROOT+r'\trends\temporary_files'
        fs=FileSystemStorage(location=location)
        file_2G_sheet=fs.save(site_list_4G.name,site_list_4G)
        file_2G_path=fs.path(file_2G_sheet)
        df_2G_site=pd.read_excel(file_2G_path)
        print(df_2G_site)
        os.remove(path=file_2G_path)     


    door_path=os.path.join(MEDIA_ROOT,'trends','pb','pb_zte')
  
    
    df_raw_kpi_4G['Short name']=df_raw_kpi_4G['Short name'].fillna(method='ffill')
 
    df_raw_kpi_4G.rename(columns={'Unnamed: 1':"Date"},inplace=True)
    df_raw_kpi_4G['MV_DL User Throughput_Kbps [CDBH]']=(df_raw_kpi_4G['MV_DL User Throughput_Kbps [CDBH]']/1024)

    df_raw_kpi_4G.rename(columns={'MV_DL User Throughput_Kbps [CDBH]':'MV_DL User Throughput_Mbps [CDBH]'},inplace=True)
    df_raw_kpi_4G['MV_UL User Throughput_Kbps [CDBH]']=(df_raw_kpi_4G['MV_UL User Throughput_Kbps [CDBH]']/1024)
    df_raw_kpi_4G.rename(columns={'MV_UL User Throughput_Kbps [CDBH]':'MV_UL User Throughput_Mbps [CDBH]'},inplace=True)


    print(df_raw_kpi_4G)

    split=[]
    split1=[]
    techlist=[]
    remove=[]
    for cell1 in df_raw_kpi_4G['Short name']:
        if ("FDD" in cell1 or "TDD" in cell1): 
            slice=cell1[:-4] 
            remove.append(slice)
        else:
            slice=cell1
            remove.append(slice)
    df_raw_kpi_4G.insert(1,"short",remove) 
    for cell in df_raw_kpi_4G['short']:
    
        if('_' in cell):
            site_id=cell.split('_')[-2][:-1]
            split.append(site_id)
        else:
            site_id=cell[:-1]
            split.append(site_id)  
            
        if('_' in cell):
            cell_id=cell.split('_')[-2]
            split1.append(cell_id)
        else:
            cell_id=cell[:-1] 
            split1.append(cell_id)
            
        if('_F1_' in cell or '_F3_' in cell or '_F8_' in cell or  '_T1' in cell or '_T2_' in cell ):  
            if('_F1_' in cell):
                tech='L2100' 
            if('_F3_' in cell):
                tech='L1800' 
            if('_F8_' in cell):
                tech='L900'       
            if('_T1_' in cell or '_T2_' in cell):
                tech='L2300' 
            techlist.append(tech)
        else:
            cell=site_id
            techlist.append(tech)
   
      
    df_raw_kpi_4G.insert(0,'Site ID',split)
    df_raw_kpi_4G.insert(2,'cell_id',split1)
    df_raw_kpi_4G.insert(5,'Tech',techlist)
    # df.insert(6,"short",remove) 
    df_raw_kpi_4G.drop('Short name',inplace=True,axis=1)
    df_raw_kpi_4G.rename(columns={"Short name" :'short' } ,inplace = True )

    # df_raw_kpi_4G.rename(columns={"Short name" :"Shortname" } ,inplace = True )
    ecgi=[]
    ecgi1=[]
    for cell1 in df_raw_kpi_4G['4G_ECGI'] :
        if('-' in cell1):
            lnbts_id=cell1.split('-')[-2]
            ecgi.append(lnbts_id)  
        else:
            lnbts_id=cell1[:-1]
            ecgi.append(lnbts_id) 
            
        if('-' in cell1):
            lncell_id=cell1.split('-')[-1]
            ecgi1.append(lncell_id)
        else:
            lncell_id=cell1[-1]
            ecgi1.append(lncell_id)  
    df_raw_kpi_4G.insert(3,'lnbts id',ecgi)
    df_raw_kpi_4G.insert(4,'lncell_id',ecgi1)  


    df_raw_kpi_4G['Site name']=df_raw_kpi_4G['Site ID'].astype(str)+'('+ df_raw_kpi_4G['lnbts id']+')'

    df_raw_kpi_4G.rename(columns={"lnbts id" :"lnbts_id" } ,inplace = True )

    df_raw_kpi_4G.fillna(value=0,inplace=True) 
    PsOs_Path1=os.path.join(door_path,'process output','desired input.xlsx')                    
    df_raw_kpi_4G.to_excel(PsOs_Path1,index=False)
    
    excel_file_1 = PsOs_Path1
    # PsOs_Path2=os.path.join(door_path,'project file','site.xlsx')
    # excel_file_2 = PsOs_Path2

    df1 = pd.read_excel(excel_file_1)
    # df2 = pd.read_excel(excel_file_2)
    
    df1.rename(columns={"lnbts id": "lnbts_id"}, inplace=True)
    

    kpi=['MV_RRC Setup Success Rate [CDBH]','ERAB Setup Success Rate [CDBH]','MV_PS Drop Call Rate % [CDBH]','MV_UL User Throughput_Mbps [CDBH]',
        'MV_DL User Throughput_Mbps [CDBH]','MV_E-UTRAN Average CQI [CDBH]',
        'MV_PS handover success rate [LTE Intra System] [CDBH]',
        'MV_PS handover success rate [LTE Inter System] [CDBH]','MV_CSFB Redirection Success Rate [CDBH]',
        'Paging record discarded At eNodeB [CDBH]',
        'MV_Average number of used DL PRBs [CDBH]','MV_VoLTE ERAB Setup Success Rate [CBBH]',
        'MV_VoLTE DCR [CBBH]','MV_VoLTE Packet Loss DL [CBBH]','MV_VoLTE Packet Loss UL [CBBH]','VoLTE SRVCC SR [CBBH]',
        'MV_PUSCH SINR [CBBH]','MV_VoLTE IntraF HOSR Exec [CBBH]','MV_VoLTE InterF HOSR Exec [CBBH]',
        'MV_VoLTE SRVCC Per Call Rate [CBBH]',
        'UL RSSI','MV_4G Data Volume_GB','MV_VoLTE Traffic','PS handover success rate_NOM [LTE Intra System] [CDBH]',
        'PS handover success rate_DENOM [LTE Intra System] [CDBH]','PS handover success rate_NOM [LTE Inter System] [CDBH]','PS handover success rate_DENOM [LTE Inter System] [CDBH]','VoLTE IntraF HOSR Exec_NOM [CBBH]',
        'VoLTE IntraF HOSR Exec_DENOM [CBBH]','VoLTE InterF HOSR Exec_NOM [CBBH]',
        'VoLTE InterF HOSR Exec_DENOM [CBBH]','MV_VoLTE SRVCC Per Call Rate_Nom [CBBH]',
        'MV_VoLTE SRVCC Per Call Rate_Denom [CBBH]','MV_Average number of used UL PRBs [CDBH]',
        'TA Sampls > 1.5 Km % [CDBH]','VoLTE Packet Loss DL_Nom [CBBH]',
        'VoLTE Packet Loss UL_Nom [CBBH]']
    
    df_filter = df1[(df1.lnbts_id.isin(list(df_site_4G['2G ID'])))]

    print(df_filter)
    df_filter1=df_filter.round(2)
    PsOs_Filter=os.path.join(door_path,'process output','filtered_df_1.xlsx')
    df_filter1.to_excel(PsOs_Filter,index=False)

    df1 = pd.read_excel(PsOs_Filter)
       
    df_pivot = df1.pivot_table(values=kpi, columns='Date', index=['short','Site ID','cell_id','lnbts_id','lncell_id','Site name','Tech','4G_ECGI'])
    
    PsOs_Pivot=os.path.join(door_path,'process output','pivot.xlsx')
    df_pivot.to_excel(PsOs_Pivot)

    str_date=request.POST.get("offered_date")
    date1=datetime.datetime.strptime(str_date,'%Y-%m-%d')
    d1=date1-timedelta(1)
    d2=date1-timedelta(2)
    d3=date1-timedelta(3)
    d4=date1-timedelta(4)
    d5=date1-timedelta(5)
    cl=[d1,d2,d3,d4,d5]
    index=df_pivot.index
    
    STR=os.path.join(door_path,'template','punjabtemplate.xlsx')
    wb=openpyxl.load_workbook(STR)
    ws=wb['ALL TECH KPI']
################################################
    def overwrite(kpi_name,coln1,coln2,coln3,coln4,coln5):
        
        dr=df_pivot[kpi_name]
        li=dr.columns.to_list()
        col1=dr[li[0]].to_list()
        col2=dr[li[1]].to_list()
        col3=dr[li[2]].to_list()
        col4=dr[li[3]].to_list()
        col5=dr[li[4]].to_list()


        ws[coln1+"4"].value=cl[4]
        ws[coln2+"4"].value=cl[3]
        ws[coln3+"4"].value=cl[2]
        ws[coln4+"4"].value=cl[1]
        ws[coln5+"4"].value=cl[0]



        for i,value in enumerate(index):
                j=i+5
                ws['B'+str(j)].value='PB'
            
                # ws['L'+str(j)].value=date1
                # ws['P'+str(j)].value='Site Type'
                

                ws['M'+str(j)].value='DONE'
                ws['A'+str(j)].value=index[i][7]
                ws['C'+str(j)].value=index[i][1]
                ws['D'+str(j)].value=index[i][2]
                ws['E'+str(j)].value=index[i][3]
                ws['F'+str(j)].value=index[i][4]
                ws['G'+str(j)].value=index[i][1]
                ws['H'+str(j)].value=index[i][0]
                ws['I'+str(j)].value=index[i][5]
                ws['J'+str(j)].value=index[i][6]
                
                ws[coln1+str(j)].value=col1[i]
                ws[coln2+str(j)].value=col2[i]
                ws[coln3+str(j)].value=col3[i]
                ws[coln4+str(j)].value=col4[i]
                ws[coln5+str(j)].value=col5[i]
    
    for kpi_name in kpi:
        if(kpi_name=='MV_RRC Setup Success Rate [CDBH]'):
            overwrite(kpi_name,'Q','R','S','T','U')   
        if(kpi_name=='ERAB Setup Success Rate [CDBH]'):
            overwrite(kpi_name,'X','Y','Z','AA','AB')
            
        if(kpi_name=='MV_PS Drop Call Rate % [CDBH]'):
            overwrite(kpi_name,'AE','AF','AG','AH','AI')
        if(kpi_name=='MV_DL User Throughput_Mbps [CDBH]'):
            overwrite(kpi_name,'AL','AM','AN','AO','AP')
        if(kpi_name=='MV_UL User Throughput_Mbps [CDBH]'):
            overwrite(kpi_name,'AT','AU','AV','AW','AX')
        if(kpi_name=='MV_E-UTRAN Average CQI [CDBH]'):
            overwrite(kpi_name,'BA','BB','BC','BD','BE')
        if(kpi_name=='MV_PS handover success rate [LTE Intra System] [CDBH]'):
            overwrite(kpi_name,'BH','BI','BJ','BK','BL')
        if(kpi_name=='MV_PS handover success rate [LTE Inter System] [CDBH]'):
            overwrite(kpi_name,'BO','BP','BQ','BR','BS')
        if(kpi_name=='MV_CSFB Redirection Success Rate [CDBH]'):
            overwrite(kpi_name,'BV','BW','BX','BY','BZ')
        if(kpi_name=='Paging record discarded At eNodeB [CDBH]'):
            overwrite(kpi_name,'CC','CD','CE','CF','CG')
        if(kpi_name=='MV_Average number of used DL PRBs [CDBH]'):
            overwrite(kpi_name,'CJ','CK','CL','CM','CN')  
        if(kpi_name=='MV_VoLTE ERAB Setup Success Rate [CBBH]'):
            overwrite(kpi_name,'CQ','CR','CS','CT','CU')  
        if(kpi_name=='MV_VoLTE DCR [CBBH]'):
            overwrite(kpi_name,'CX','CY','CZ','DA','DB') 
        if(kpi_name=='MV_VoLTE Packet Loss DL [CBBH]'):
            overwrite(kpi_name,'DE','DF','DG','DH','DI') 
        if(kpi_name=='MV_VoLTE Packet Loss UL [CBBH]'):
            overwrite(kpi_name,'DL','DM','DN','DO','DP') 
        if(kpi_name=='VoLTE SRVCC SR [CBBH]'):
            overwrite(kpi_name,'DS','DT','DU','DV','DW') 
        if(kpi_name=='MV_PUSCH SINR [CBBH]'):
            overwrite(kpi_name,'DZ','EA','EB','EC','ED') 
        if(kpi_name=='MV_VoLTE IntraF HOSR Exec [CBBH]'):
            overwrite(kpi_name,'EG','EH','EI','EJ','EK') 
        if(kpi_name=='MV_VoLTE InterF HOSR Exec [CBBH]'):
            overwrite(kpi_name,'EN','EO','EP','EQ','ER') 
        if(kpi_name=='MV_VoLTE SRVCC Per Call Rate [CBBH]'):
            overwrite(kpi_name,'EU','EV','EW','EX','EY') 
        if(kpi_name=='UL RSSI'):
            overwrite(kpi_name,'FB','FC','FD','FE','FF') 
        if(kpi_name=='MV_4G Data Volume_GB'):
            overwrite(kpi_name,'FI','FJ','FK','FL','FM') 
        if(kpi_name=='MV_VoLTE Traffic'):
            overwrite(kpi_name,'FP','FQ','FR','FS','FT') 
        if(kpi_name=='PS handover success rate_NOM [LTE Intra System] [CDBH]'):
            overwrite(kpi_name,'FV','FW','FX','FY','FZ')
        if(kpi_name=='PS handover success rate_DENOM [LTE Intra System] [CDBH]'):
            overwrite(kpi_name,'GA','GB','GC','GD','GE') 
        if(kpi_name=='PS handover success rate_NOM [LTE Inter System] [CDBH]'):
            overwrite(kpi_name,'GL','GM','GN','GO','GP') 
        if(kpi_name=='PS handover success rate_DENOM [LTE Inter System] [CDBH]'):
            overwrite(kpi_name,'GQ','GR','GS','GT','GU') 
        if(kpi_name=='VoLTE IntraF HOSR Exec_NOM [CBBH]'):
            overwrite(kpi_name,'HB','HC','HD','HE','HF') 
        if(kpi_name=='VoLTE IntraF HOSR Exec_DENOM [CBBH]'):
            overwrite(kpi_name,'HG','HH','HI','HJ','HK')
        if(kpi_name=='VoLTE InterF HOSR Exec_NOM [CBBH]'):
            overwrite(kpi_name,'HR','HS','HT','HU','HV') 
        if(kpi_name=='VoLTE InterF HOSR Exec_DENOM [CBBH]'):
            overwrite(kpi_name,'HW','HX','HY','HZ','IA') 
        if(kpi_name=='MV_VoLTE SRVCC Per Call Rate_Nom [CBBH]'):
            overwrite(kpi_name,'IH','II','IJ','IK','IL') 
        if(kpi_name=='MV_VoLTE SRVCC Per Call Rate_Denom [CBBH]'):
            overwrite(kpi_name,'IM','IN','IO','IP','IQ') 
        if(kpi_name=='MV_Average number of used UL PRBs [CDBH]'):
            overwrite(kpi_name,'IX','IY','IZ','JA','JB')
        if(kpi_name=='TA Sampls > 1.5 Km % [CDBH]'):
            overwrite(kpi_name,'JN','JO','JP','JQ','JR')
        if(kpi_name=='VoLTE Packet Loss DL_Nom [CBBH]'):
            overwrite(kpi_name,'JT','JU','JV','JW','JX')  
        if(kpi_name=='VoLTE Packet Loss UL_Nom [CBBH]'):
            overwrite(kpi_name,'KA','KB','KC','KD','KE') 
            
################# GSM FILTE PIVOT#####################    
        
    door_path=os.path.join(MEDIA_ROOT,'trends','pb','pb_zte')        
    # df_raw_kpi_2G=pd.read_excel('actual input/PB 2G TOOL KPI.xlsx')
    # df_raw_kpi_2G=pd.read_excel(kpi_file["text"])
    
    df_raw_kpi_2G['Short name']=df_raw_kpi_2G['Short name'].fillna(method='ffill')
  
    
    df_raw_kpi_2G.columns.values[1]='Date'

    a=[]
    b=[]
    for cell1 in df_raw_kpi_2G['Short name']:
        if ('-' in cell1):
            siteid=cell1.split('-')[-1][1:-1]
            a.append(siteid)
        else:
            siteid=cell1[:-1]  
            a.append(siteid) 
            
        if ('-' in cell1):
            cellid=cell1.split('-')[-1] 
            b.append(cellid)
        else:
            cellid=cell1[:-1]
            b.append(cellid) 
    # for cell1 in df_raw_kpi_2G['Short name']:
    
    #     cellid=cell1[:-1]
    #     b.append(cellid)         
    df_raw_kpi_2G.insert(0,'Site ID',b)
    df_raw_kpi_2G.insert(3,'Cell ID',b) 
    
    c=[]
    for cell2 in df_raw_kpi_2G['2G CGI']:
        if('-' in cell2):
            cgi=cell2.split('-')[-1]
            c.append(cgi)
    df_raw_kpi_2G.insert(2,'cgisplit',c)    

    # df_raw_kpi_2G.rename(columns={"Short name" :"Shortname"} ,inplace = True)
    df_raw_kpi_2G.fillna(value=0,inplace=True)                     
    g2_path2=os.path.join(door_path,'process output','2Gfill.xlsx')
    df_raw_kpi_2G.to_excel(g2_path2, index=False)
    excel_1 = g2_path2
    
    excel_2G_raw= pd.read_excel(excel_1)


    excel_2G_raw.rename(columns={"cgisplit": "cgi_split"}, inplace=True)
    df_2G_site.rename(columns={"cgisplit": "cgi_split"}, inplace=True)
    gsm=['SDCCH Blocking Rate [BBH]','SDCCH Drop Call Rate [BBH]','TCH Blocking Rate [BBH]','TCH Drop Call Rate [BBH]',
            'Handover Success Rate [BBH]','RX Quality [BBH]','Total Voice Traffic','Handover Success Rate_Nom [BBH]',
            'Handover Success Rate_Denom [BBH]',
            'Drop Call Rate_Nom [BBH]','Drop Call Rate_Denom [BBH]','Number of TRX [BBH]',
            'ICM%[BBH]','Cell Downtime [sec] [BBH]','TCH Utilization [BBH]'] 

    filter_df=excel_2G_raw[(excel_2G_raw.cgi_split.isin(list(df_2G_site['2G ID'])))]
    print(filter_df)
    g2_filter=os.path.join(door_path,'process output','2Gfilter.xlsx')
    filter_df.to_excel(g2_filter, index=False)
 
    df1=pd.read_excel(g2_filter)
   

    
    # df_pivot1=df_raw_kpi_2G.pivot_table(values=gsm,columns='Date',index=['Cell ID','Site_ID','2G CGI'])
    df_pivot1=df1.pivot_table(values=gsm,columns='Date',index=['Site ID','2G CGI','Short name'])
    g2_pivot=os.path.join(door_path,'process output','2gpivot.xlsx')
    df_pivot1.to_excel(g2_pivot)

    
    index=df_pivot1.index
    def overwrite(gsm_name,coln1,coln2,coln3,coln4,coln5,ws1):
        dr=df_pivot1[gsm_name]
        li=dr.columns
        col1=dr[li[0]].to_list()
        col2=dr[li[1]].to_list()
        col3=dr[li[2]].to_list()
        col4=dr[li[3]].to_list()
        col5=dr[li[4]].to_list()


        ws1[coln1+"4"].value=cl[4]
        ws1[coln2+"4"].value=cl[3]
        ws1[coln3+"4"].value=cl[2]
        ws1[coln4+"4"].value=cl[1]
        ws1[coln5+"4"].value=cl[0]
    
    
        for i,value in enumerate(index):
                j=i+5
                ws1['F'+str(j)].value='GSM'
                ws1['L'+str(j)].value='Site'
                ws1['B'+str(j)].value='PB'
                
                # ws1['H'+str(j)].value=date1
                ws1['D'+str(j)].value=index[i][2]
                ws1['C'+str(j)].value=index[i][0]
                ws1['A'+str(j)].value=index[i][1]
                ws1[coln1+str(j)].value=col1[i]
                ws1[coln2+str(j)].value=col2[i]
                ws1[coln3+str(j)].value=col3[i]
                ws1[coln4+str(j)].value=col4[i]
                ws1[coln5+str(j)].value=col5[i]        

    g2_ws=wb['2G KPI']        
    for gsm_name in gsm:
        if(gsm_name=='SDCCH Blocking Rate [BBH]'):
            overwrite(gsm_name,'M','N','O','P','Q',g2_ws)

        if(gsm_name=='SDCCH Drop Call Rate [BBH]'):
            overwrite(gsm_name,'T','U','V','W','X',g2_ws) 
        if(gsm_name=='SDCCH Drop Call Rate [BBH]'):
            overwrite(gsm_name,'T','U','V','W','X',g2_ws) 
        if(gsm_name=='TCH Blocking Rate [BBH]'):
            overwrite(gsm_name,'AA','AB','AC','AD','AE',g2_ws) 
        if(gsm_name=='TCH Drop Call Rate [BBH]'):
            overwrite(gsm_name,'AH','AI','AJ','AK','AL',g2_ws) 
        if(gsm_name=='Handover Success Rate [BBH]'):
            overwrite(gsm_name,'AP','AQ','AR','AS','AT',g2_ws) 
            
        if(gsm_name=='RX Quality [BBH]'):
            overwrite(gsm_name,'AW','AX','AY','AZ','BA',g2_ws) 
        if(gsm_name=='Total Voice Traffic'):
            overwrite(gsm_name,'BK','BL','BM','BN','BO',g2_ws) 
        if(gsm_name=='Handover Success Rate_Nom [BBH]'):
            overwrite(gsm_name,'BQ','BR','BS','BT','BU',g2_ws)    
        if(gsm_name=='Handover Success Rate_Denom [BBH]'):
            overwrite(gsm_name,'BV','BW','BX','BY','BZ',g2_ws) 
        if(gsm_name=='Drop Call Rate_Nom [BBH]'):
            overwrite(gsm_name,'CG','CH','CI','CJ','CK',g2_ws) 
        if(gsm_name=='Drop Call Rate_Denom [BBH]'):
            overwrite(gsm_name,'CL','CM','CN','CO','CP',g2_ws) 
        if(gsm_name=='Drop Call Rate_Nom [BBH]'):
            overwrite(gsm_name,'CQ','CR','CS','CT','CU',g2_ws) 
        if(gsm_name=='Number of TRX [BBH]'):
            overwrite(gsm_name,'DM','DN','DO','DP','DQ',g2_ws) 
        if(gsm_name=='ICM%[BBH]'):
            overwrite(gsm_name,'DS','DT','DU','DV','DW',g2_ws) 
        if(gsm_name=='Cell Downtime [sec] [BBH]'):
            overwrite(gsm_name,'DY','DZ','EA','EB','EC',g2_ws) 
        if(gsm_name=='TCH Utilization [BBH]'):
            overwrite(gsm_name,'EE','EF','EG','EH','EI',g2_ws)         
    save_output=os.path.join(door_path,'output','Punjab_trend_output.xlsx')
    wb.save(save_output) 

    download_path=os.path.join(MEDIA_URL,'trends','pb','pb_zte','output','Punjab_trend_output.xlsx') 
    return Response({'status':True,'message':'successfully','Download_url':download_path})

##################################### nokia pb ###################################################

@api_view(["POST"])
def old_pb_nok_smallcell_trend(request):
    raw_kpi_4G=request.FILES['raw_kpi_4G'] if 'raw_kpi_4G' in request.FILES else None
    if raw_kpi_4G:
        location=MEDIA_ROOT +r'\trends\temporary_files'
        fs=FileSystemStorage(location=location)
        file=fs.save(raw_kpi_4G.name,raw_kpi_4G)
        file_path=fs.path(file)
        df_raw_kpi_4G=pd.read_excel(file_path)
        print(df_raw_kpi_4G)
        os.remove(path=file_path)

    site_list_4G=request.FILES['site_list_4G'] if 'site_list_4G' in request.FILES else None
    if site_list_4G:
        location=MEDIA_ROOT+r'\trends\temporary_files'
        fs=FileSystemStorage(location=location)
        file_sheet=fs.save(site_list_4G.name,site_list_4G)
        file_path=fs.path(file_sheet)
        df_site_4G=pd.read_excel(file_path)
        print(df_site_4G)
        os.remove(path=file_path)


    door_path=os.path.join(MEDIA_ROOT,'trends','pb','pb_nok')
    
    temp=os.path.join(door_path,'template','Small Cell KPI Trend PB.xlsx')
    STR=temp
    wb=openpyxl.load_workbook(STR)
    # ws=wb['ALL TECH KPI']
    ws=wb.active
    
    df_raw_kpi_4G["Short name"]=df_raw_kpi_4G["Short name"].fillna(method="ffill")
    df_raw_kpi_4G.columns.values[1]="Date"
    
    df_raw_kpi_4G.fillna(value=0,inplace=True) 
    split=[]
    split1=[]
    techlist=[]
    for cell1 in df_raw_kpi_4G['Short name']:
        if('_' in cell1):
            site_id=cell1.split('_')[-2][:-1]
            split.append(site_id)
            
        else:
            site_id=cell1[:-1]  
            split.append(site_id)
        if('_' in cell1):
            site_id1=cell1.split('_')[-2]
            split1.append(site_id1)
        else:
            site_id1=cell1[:-1]
            split1.append(site_id1)
            
        if('_F1_' in cell1 or '_F3_' in cell1 or '_F8_' in cell1 or  '_T1' in cell1 or '_T2_' in cell1 ):  
         
            if('_F1_' in cell1):
                tech='L2100' 
            if('_F3_' in cell1):
                tech='L1800' 
            if('_F8_' in cell1):
                tech='L900'       
            if('_T1_' in cell1 or '_T2_' in cell1):
                tech='L2300' 
            techlist.append(tech)
        else:
            tech=site_id1
            techlist.append(tech)     
              
    df_raw_kpi_4G.insert(0,"site_id",split)   
    df_raw_kpi_4G.insert(2,'cellid',split1)
    df_raw_kpi_4G.insert(4,'tech',techlist)
    
    
    
    ecgi=[]
    ecgi1=[]
    for cell2 in df_raw_kpi_4G["4G_ECGI"]:
        if('-' in cell2):
            ecgi_id=cell2.split('-')[-2]
            ecgi.append(ecgi_id)
        else:
            ecgi_id=cell2[:-1] 
            ecgi.append(ecgi_id)
        if('-' in cell2):
            ecgi_id1=cell2.split('-')[-1]
            ecgi1.append(ecgi_id1)
        else:
            ecgi_id1=cell2[-1]
            ecgi1.append(ecgi_id1)         
    df_raw_kpi_4G.insert(3,'ecgi',ecgi)
    df_raw_kpi_4G.insert(5,'ecgicell',ecgi1) 
    
    df_raw_kpi_4G['ECGI ID']=df_raw_kpi_4G['ecgi'].astype(str)+'_'+ df_raw_kpi_4G['ecgicell']
    df_raw_kpi_4G['LNCEL NAME']=df_raw_kpi_4G['Short name'].astype(str)+'(id:'+ df_raw_kpi_4G['ecgicell']+')' 
    df_raw_kpi_4G['LNBTS NAME']=df_raw_kpi_4G['Short name'].astype(str)+'(id:'+ df_raw_kpi_4G['ecgi']+')' 
   
    # df.loc[df['name']]='mrbts'
    df_raw_kpi_4G['mrbtsname']='mrbts'
    df_raw_kpi_4G['MRBTS NAME']=df_raw_kpi_4G['mrbtsname'].astype(str)+'-'+ df_raw_kpi_4G['LNBTS NAME']+')' 
    # print(df.column_name)
    
     
    df_raw_kpi_4G["Short name"] =df_raw_kpi_4G["Short name"].apply(lambda x: x.strip())
    PsOs_path1=os.path.join(door_path,'process output','desiredsmallcell input.xlsx')
    
    df_raw_kpi_4G.to_excel(PsOs_path1,index=False) 
    
    excel_file1=PsOs_path1 
    # excel_file2='project file/site.xlsx'
    
    df1=pd.read_excel(excel_file1)
    # df2=pd.read_excel(excel_file2)
    
    kpi=['RRC Setup Success Rate [CDBH]','ERAB Setup Success Rate [CDBH]','PS Drop Call Rate % [CDBH]',
         'MV_DL User Throughput_Kbps [CDBH]','MV_UL User Throughput_Kbps [CDBH]','PS handover success rate [LTE Intra System] [CDBH]',
         'PS handover success rate [LTE Inter System] [CDBH]','CSFB Redirection Success Rate [CDBH]','VoLTE ERAB Setup Success Rate [CBBH]',
         'VoLTE Drop Call Rate [CBBH]]','MV_VoLTE Packet Loss DL [CBBH]','MV_VoLTE Packet Loss UL [CBBH]','VoLTE SRVCC SR [CBBH]','VoLTE Intra-LTE Handover Success Ratio [CBBH]',
         'VoLTE Inter-Frequency Handover Success Ratio [CBBH]','E-UTRAN Average CQI [CDBH]','UL RSSI [CBBH]',
         'MV_4G Data Volume_GB','VoLTE Traffic','MV_Average number of used DL PRBs [CDBH]']
    
    df_filter = df1[(df1.ecgi.isin(list(df_site_4G['2G ID'])))]
    print(df_filter)
    PsOs_filter=os.path.join(door_path,'process output','filteredsmallcell.xlsx')
    df_filter.to_excel(PsOs_filter, index=False)
    
    df1 = pd.read_excel(PsOs_filter)
       
    df_pivot = df1.pivot_table(columns='Date', index=['site_id','cellid','tech','LNCEL NAME','ECGI ID','ecgi','LNBTS NAME','MRBTS NAME'])
     
    PsOs_pivot=os.path.join(door_path,'process output','pivotsmallcell.xlsx') 
    df_pivot.to_excel(PsOs_pivot)
    
    
    # STR='template/Small Cell KPI Trend PB.xlsx'
    # wb=openpyxl.load_workbook(STR)
    # # ws=wb['ALL TECH KPI']
    # ws=wb.active
    
    
    
    alpha = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    def num_hash(num):
            if num < 26:
                return alpha[num-1]
            else:
                q, r = num//26, num % 26
                if r == 0:
                    if q == 1:
                        return alpha[r-1]
                    else:
                        return num_hash(q-1) + alpha[r-1]
                else:
                    return num_hash(q) + alpha[r-1]
            
            
    def titleToNumber(s):
	# This process is similar to binary-to-
	# decimal conversion
        result = 0
        for B in range(len(s)):
            result *= 26
            result += ord(s[B]) - ord('A') + 1
        return result
    
    str_date=request.POST.get('offered_date')

    date1=datetime.datetime.strptime(str_date,'%Y-%m-%d')
    d1=date1-timedelta(1)
    d2=date1-timedelta(2)
    d3=date1-timedelta(3)
    d4=date1-timedelta(4)
    d5=date1-timedelta(5)
    cl=[d1,d2,d3,d4,d5]
   
    def overwrite(kpi_name,coln1):
        coln2=num_hash(titleToNumber(coln1)+1)
        coln3=num_hash(titleToNumber(coln1)+2)
        coln4=num_hash(titleToNumber(coln1)+3)
        coln5=num_hash(titleToNumber(coln1)+4)
        print(kpi_name)
        index=df_pivot.index
        print('donnnnnnnnnnnnnnnnne')
        print(len(index))
        dr=df_pivot[kpi_name]
        
        li=dr.columns
        col1=dr[li[0]].to_list()
        col2=dr[li[1]].to_list()
        col3=dr[li[2]].to_list()
        col4=dr[li[3]].to_list()
        col5=dr[li[4]].to_list()


        ws[coln1+"4"].value=cl[4]
        ws[coln2+"4"].value=cl[3]
        ws[coln3+"4"].value=cl[2]
        ws[coln4+"4"].value=cl[1]
        ws[coln5+"4"].value=cl[0]



        for i,value in enumerate(index):
                j=i+5
                ws['A'+str(j)].value='PB'
                # ws['M'+str(j)].value='ERICSSON'
                ws['N'+str(j)].value=date1

                ws['L'+str(j)].value='DONE'
                ws['B'+str(j)].value=index[i][0]
                ws['C'+str(j)].value=index[i][1]
                ws['D'+str(j)].value=index[i][5]
                ws['G'+str(j)].value=index[i][7]
                
                ws['H'+str(j)].value=index[i][6]
                
                ws['I'+str(j)].value=index[i][3]
                ws['J'+str(j)].value=index[i][4]
                ws['L'+str(j)].value=index[i][2]     
                ws[coln1+str(j)].value=col1[i]
                ws[coln2+str(j)].value=col2[i]
                ws[coln3+str(j)].value=col3[i]
                ws[coln4+str(j)].value=col4[i]
                ws[coln5+str(j)].value=col5[i]
    for kpi_name in kpi:
        if(kpi_name=='RRC Setup Success Rate [CDBH]'):
            overwrite(kpi_name,'R') 
        if(kpi_name=='ERAB Setup Success Rate [CDBH]'):
            overwrite(kpi_name,'X') 
        if(kpi_name=='PS Drop Call Rate % [CDBH]'):
            overwrite(kpi_name,'AD') 
        if(kpi_name=='MV_DL User Throughput_Kbps [CDBH]'):
            overwrite(kpi_name,'AJ')     
        if(kpi_name=='MV_UL User Throughput_Kbps [CDBH]'):
            overwrite(kpi_name,'AR')     
        if(kpi_name=='PS handover success rate [LTE Intra System] [CDBH]'):
            overwrite(kpi_name,'AX')     
        if(kpi_name=='PS handover success rate [LTE Inter System] [CDBH]'):
            overwrite(kpi_name,'BD')     
        if(kpi_name=='CSFB Redirection Success Rate [CDBH]'):
            overwrite(kpi_name,'BJ')     
        if(kpi_name=='VoLTE ERAB Setup Success Rate [CBBH]'):
            overwrite(kpi_name,'BP')     
        if(kpi_name=='VoLTE Drop Call Rate [CBBH]]'):
            overwrite(kpi_name,'BV') 
        if(kpi_name=='MV_VoLTE Packet Loss DL [CBBH]'):
            overwrite(kpi_name,'CB') 
        if(kpi_name=='MV_VoLTE Packet Loss UL [CBBH]'):
            overwrite(kpi_name,'CH')         
        if(kpi_name=='VoLTE SRVCC SR [CBBH]'):
            overwrite(kpi_name,'CN')     
        if(kpi_name=='VoLTE Intra-LTE Handover Success Ratio [CBBH]'):
            overwrite(kpi_name,'CT') 
        if(kpi_name=='VoLTE Inter-Frequency Handover Success Ratio [CBBH]'):
            overwrite(kpi_name,'CZ')        
        if(kpi_name=='E-UTRAN Average CQI [CDBH]'):
            overwrite(kpi_name,'DF') 
        if(kpi_name=='UL RSSI [CBBH]'):
            overwrite(kpi_name,'DN')     
        if(kpi_name=='MV_4G Data Volume_GB'):
            overwrite(kpi_name,'DT') 
        if(kpi_name=='VoLTE Traffic'):
            overwrite(kpi_name,'EB')     
        if(kpi_name=='MV_Average number of used DL PRBs [CDBH]'):
            overwrite(kpi_name,'EH')         
            
    save_output=os.path.join(door_path,'output','smallcell Trend.xlsx')           
                  
   
    wb.save(save_output)      
    download_path=os.path.join(MEDIA_URL,"trends","pb","pb_nok","output","smallcell Trend.xlsx") 
    return Response({'status':True,'message':'successfully','Download_url':download_path})
################################### HPSCNOKIA ###############################################
@api_view(["POST"])
def old_pb_nok_hpsccell_trend(request):
    raw_kpi_4G=request.FILES['raw_kpi_4G'] if 'raw_kpi_4G' in request.FILES else None
    if raw_kpi_4G:
        location=MEDIA_ROOT +r'\trends\temporary_files'
        fs=FileSystemStorage(location=location)
        file=fs.save(raw_kpi_4G.name,raw_kpi_4G)
        file_path=fs.path(file)
        df_raw_kpi_4G=pd.read_excel(file_path)
        print(df_raw_kpi_4G)
        os.remove(path=file_path)

    site_list_4G=request.FILES['site_list_4G'] if 'site_list_4G' in request.FILES else None
    if site_list_4G:
        location=MEDIA_ROOT+r'\trends\temporary_files'
        fs=FileSystemStorage(location=location)
        file_sheet=fs.save(site_list_4G.name,site_list_4G)
        file_path=fs.path(file_sheet)
        df_site_4G=pd.read_excel(file_path)
        print(df_site_4G)
        os.remove(path=file_path)

    door_path=os.path.join(MEDIA_ROOT,'trends','pb','pb_nok')
    
    
    df_raw_kpi_4G["Short name"]=df_raw_kpi_4G["Short name"].fillna(method="ffill")
    # df_raw_kpi_4G.columns.values[1]="Date"
    df_raw_kpi_4G.rename(columns={'Unnamed: 1':"Date"},inplace=True)
    # df["MV_DL User Throughput_Kbps [CDBH]"]=(df["MV_DL User Throughput_Kbps [CDBH]"]/1024)
    # df.columns.values[16]="MV_DL User Throughput_Mbps [CDBH]"
    print(df_raw_kpi_4G)

    split=[]
    split1=[]
    for cell1 in df_raw_kpi_4G['Short name']:
        if('_' in cell1):
            site_id=cell1.split('_')[-2][:-1]
            split.append(site_id)
            
        else:
            site_id=cell1[:-1]  
            split.append(site_id)
            
        if('_' in cell1):
            site_id1=cell1.split('_')[-2]
            split1.append(site_id1)
        else:
            site_id1=cell1[:-1]
            split1.append(site_id1)     
    df_raw_kpi_4G.insert(0,"site_id",split) 
    df_raw_kpi_4G.insert(2,'cellid',split1)  
    
    
    
    ecgi=[]
    ecgi1=[]
    for cell2 in df_raw_kpi_4G["4G_ECGI"]:
        if('-' in cell2):
            ecgi_id=cell2.split('-')[-2]
            ecgi.append(ecgi_id)
        else:
            ecgi_id=cell2[:-1] 
            ecgi.append(ecgi_id)
            
        if('-' in cell2):
            ecgi_id1=cell2.split('-')[-1]
            ecgi1.append(ecgi_id1)
        else:
            ecgi_id1=cell2[-1]
            ecgi1.append(ecgi_id1)     
              
    df_raw_kpi_4G.insert(2,'ecgi',ecgi)
    df_raw_kpi_4G.insert(3,'ecgicell',ecgi1) 
    df_raw_kpi_4G['CELL NAME']=df_raw_kpi_4G['Short name'].astype(str)+'(id:'+ df_raw_kpi_4G['ecgicell']+')' 
    df_raw_kpi_4G['ECGI ID']=df_raw_kpi_4G['ecgi'].astype(str)+'_'+ df_raw_kpi_4G['ecgicell']  
    df_raw_kpi_4G.fillna(value=0,inplace=True)
    
    # df.groupby('site_id')['MV_4G Data Volume_GB'].sum()
    PsOs_path1=os.path.join(door_path,'process output','desirednokia input.xlsx')
    df_raw_kpi_4G.to_excel(PsOs_path1,index=False) 
    
    excel_file1=PsOs_path1
    # excel_file2='project file/site.xlsx'
    
    df1=pd.read_excel(excel_file1)

    
    kpi=['Radio NW Availability']
    kpi_payload=['MV_4G Data Volume_GB']
    df_filter=df1[(df1.ecgi.isin(list(df_site_4G['2G ID'])))]
    print(df_filter)
   
    PsOs_filter=os.path.join(door_path,'process output','filternokia.xlsx')
    df_filter.to_excel(PsOs_filter, index=False)
    print(df_filter)
    df3=pd.read_excel(PsOs_filter)
    pivot=df3.pivot_table(values=kpi,columns='Date',index=['site_id','ecgi',])
    # pivot_payload=df3.pivot_table(values=kpi_payload,columns='Date',index=['site_id','ecgi'],aggfunc=sum)
    print(pivot)
 
    PsOs_pivot=os.path.join(door_path,'process output','pivotcellavail.xlsx')
    pivot.to_excel(PsOs_pivot)
      
    STR=os.path.join(door_path,'template','HPSCTEMP.xlsx')
    wb=openpyxl.load_workbook(STR)
    # ws=wb['ALL TECH KPI']
    ws=wb['HPSC']
    
    alpha = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    def num_hash(num):
            if num < 26:
                return alpha[num-1]
            else:
                q, r = num//26, num % 26
                if r == 0:
                    if q == 1:
                        return alpha[r-1]
                    else:
                        return num_hash(q-1) + alpha[r-1]
                else:
                    return num_hash(q) + alpha[r-1]
            
            
    def titleToNumber(s):
	# This process is similar to binary-to-
	# decimal conversion
        result = 0
        for B in range(len(s)):
            result *= 26
            result += ord(s[B]) - ord('A') + 1
        return result
    
   
    str_date=request.POST.get('offered_date')
    date1=datetime.datetime.strptime(str_date,'%Y-%m-%d')
    # date1=date(2023,3,14)
    d1=date1-timedelta(1)
    d2=date1-timedelta(2)
    d3=date1-timedelta(3)
    d4=date1-timedelta(4)
    d5=date1-timedelta(5)
    cl=[d1,d2,d3,d4,d5]
    index=pivot.index
       
    def overwrite(kpi_name,coln1):
        coln2=num_hash(titleToNumber(coln1)+1)
        coln3=num_hash(titleToNumber(coln1)+2)
        coln4=num_hash(titleToNumber(coln1)+3)
        coln5=num_hash(titleToNumber(coln1)+4)
        print(kpi_name)
        
        dr=pivot[kpi_name]
        li=dr.columns.to_list()
        col1=dr[li[0]].to_list()
        col2=dr[li[1]].to_list()
        col3=dr[li[2]].to_list()
        col4=dr[li[3]].to_list()
        col5=dr[li[4]].to_list()


        ws[coln1+"2"].value=cl[4]
        ws[coln2+"2"].value=cl[3]
        ws[coln3+"2"].value=cl[2]
        ws[coln4+"2"].value=cl[1]
        ws[coln5+"2"].value=cl[0]



        for i,value in enumerate(index):
                j=i+3
                ws['B'+str(j)].value='PB'
                ws['C'+str(j)].value='NOKIA'
                
                ws['D'+str(j)].value=index[i][0]
                ws['H'+str(j)].value=index[i][1]
                
                
                ws[coln1+str(j)].value=col1[i]
                ws[coln2+str(j)].value=col2[i]
                ws[coln3+str(j)].value=col3[i]
                ws[coln4+str(j)].value=col4[i]
                ws[coln5+str(j)].value=col5[i]
    for kpi_name in kpi:
        if(kpi_name=='Radio NW Availability'):
            overwrite(kpi_name,'AF')        
################################################ PAYLOAD ###################################
    # index2=pivot_payload.index
    pivot_payload=df3.pivot_table(values=kpi_payload,columns='Date',index=['site_id','ecgi'],aggfunc=sum)
    PsOs_pivot_payload=os.path.join(door_path,'process output','pivotpayload.xlsx')
    pivot_payload.to_excel(PsOs_pivot_payload)
    
    index2=pivot_payload.index
    def overwrite(kpi_name1,coln1):
        coln2=num_hash(titleToNumber(coln1)+1)
        coln3=num_hash(titleToNumber(coln1)+2)
        coln4=num_hash(titleToNumber(coln1)+3)
        coln5=num_hash(titleToNumber(coln1)+4)
        print(kpi_name1)
        
        dr=pivot_payload[kpi_name1]
        lii=dr.columns.to_list()
        col1=dr[lii[0]].to_list()
        col2=dr[lii[1]].to_list()
        col3=dr[lii[2]].to_list()
        col4=dr[lii[3]].to_list()
        col5=dr[lii[4]].to_list()


        ws[coln1+"2"].value=cl[4]
        ws[coln2+"2"].value=cl[3]
        ws[coln3+"2"].value=cl[2]
        ws[coln4+"2"].value=cl[1]
        ws[coln5+"2"].value=cl[0]



        for i,value in enumerate(index2):
                j=i+3
                # ws['B'+str(j)].value='PB'
                # ws['H'+str(j)].value=index[i][1]
                
                # ws['D'+str(j)].value=index[i][0]
               
                
                ws[coln1+str(j)].value=col1[i]
                ws[coln2+str(j)].value=col2[i]
                ws[coln3+str(j)].value=col3[i]
                ws[coln4+str(j)].value=col4[i]
                ws[coln5+str(j)].value=col5[i]
    
    for kpi_name1 in kpi_payload:
        # if(kpi_name=='Radio NW Availability'):
        #     overwrite(kpi_name,'AF') 
            
        if(kpi_name1=='MV_4G Data Volume_GB'):
            overwrite(kpi_name1,'X')
            
            print('done')
            
            
            
            
            
            
    ############################ CELLWIE ##########################################    
    
    kpi_cell=['MV_4G Data Volume_GB','MV_DL User Throughput_Kbps [CDBH]','MV_UL User Throughput_Kbps [CDBH]',
              'MV_Average number of used DL PRBs [CDBH]','E-UTRAN Average CQI [CDBH]','VoLTE Drop Call Rate [CBBH]]',
              'RRC Setup Success Rate [CDBH]','ERAB Setup Success Rate [CDBH]','VoLTE Traffic','MV_VoLTE Packet Loss DL [CBBH]','MV_VoLTE Packet Loss UL [CBBH]',]
    
    pivot_cellwise=df3.pivot_table(columns='Date',index=['ecgi','site_id','CELL NAME','ecgicell','ECGI ID'])
    PsOs_pivot_cellwise=os.path.join(door_path,'process output','pivotcellwise.xlsx')
    pivot_cellwise.to_excel(PsOs_pivot_cellwise)
    
    
    # STR='template/HPSCTEMP.xlsx'
    # wb=openpyxl.load_workbook(STR)
    # ws=wb['ALL TECH KPI']
    ws=wb['CELLWISE']
    
    
    index3=pivot_cellwise.index
    def overwrite(kpi_name2,coln1):
        coln2=num_hash(titleToNumber(coln1)+1)
        coln3=num_hash(titleToNumber(coln1)+2)
        coln4=num_hash(titleToNumber(coln1)+3)
        coln5=num_hash(titleToNumber(coln1)+4)
        print(kpi_name2)
        
        dr=pivot_cellwise[kpi_name2]
        lii=dr.columns.to_list()
        col1=dr[lii[0]].to_list()
        col2=dr[lii[1]].to_list()
        col3=dr[lii[2]].to_list()
        col4=dr[lii[3]].to_list()
        col5=dr[lii[4]].to_list()


        ws[coln1+"2"].value=cl[4]
        ws[coln2+"2"].value=cl[3]
        ws[coln3+"2"].value=cl[2]
        ws[coln4+"2"].value=cl[1]
        ws[coln5+"2"].value=cl[0]



        for i,value in enumerate(index3):
                j=i+3
                ws['B'+str(j)].value='HPSC'
                ws['D'+str(j)].value='ULS'
                ws['K'+str(j)].value=index3[i][0]
                ws['I'+str(j)].value=index3[i][4]
                ws['G'+str(j)].value=index3[i][3]
               
                ws['C'+str(j)].value=index3[i][1]
                ws['F'+str(j)].value=index3[i][2]
               
               
                
                ws[coln1+str(j)].value=col1[i]
                ws[coln2+str(j)].value=col2[i]
                ws[coln3+str(j)].value=col3[i]
                ws[coln4+str(j)].value=col4[i]
                ws[coln5+str(j)].value=col5[i]
    
    for kpi_name2 in kpi_cell:
        if(kpi_name2=='MV_4G Data Volume_GB'):
            overwrite(kpi_name2,'IU') 
            
        if(kpi_name2=='MV_DL User Throughput_Kbps [CDBH]'):
            overwrite(kpi_name2,'SN')
            
        if(kpi_name2=='MV_UL User Throughput_Kbps [CDBH]'):
            overwrite(kpi_name2,'ACG') 
            
        if(kpi_name2=='MV_Average number of used DL PRBs [CDBH]'):
            overwrite(kpi_name2,'ALZ') 
            
        if(kpi_name2=='E-UTRAN Average CQI [CDBH]'):
            overwrite(kpi_name2,'AVS') 
        if(kpi_name2=='VoLTE Drop Call Rate [CBBH]]'):
            overwrite(kpi_name2,'BET') 
        if(kpi_name2=='RRC Setup Success Rate [CDBH]'):
            overwrite(kpi_name2,'BHM')  
        if(kpi_name2=='ERAB Setup Success Rate [CDBH]'):
            overwrite(kpi_name2,'BKF') 
        if(kpi_name2=='VoLTE Traffic'):
            overwrite(kpi_name2,'BMY') 
        if(kpi_name2=='MV_VoLTE Packet Loss DL [CBBH]'):
            overwrite(kpi_name2,'BPR') 
        if(kpi_name2=='MV_VoLTE Packet Loss UL [CBBH]'):
            overwrite(kpi_name2,'BSK')                      
                        
            
            print('done')  
    save_path=os.path.join(door_path,'output','HPSCNOKIA.xlsx')                  
    wb.save(save_path)
    download_path=os.path.join(MEDIA_URL,"trends","pb","pb_nok","output","HPSCNOKIA.xlsx") 
    return Response({'status':True,'message':'successfully','Download_url':download_path})

